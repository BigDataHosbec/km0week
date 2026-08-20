#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Km0 Week — generador de los descargables
========================================
Construye TODO el material que se ofrece en «Descargas» y en «Sala de prensa»
a partir de los datos reales del sitio (assets/js/data-alojamientos.js).

    python3 _build/descargables.py

Sale en   descargas/   y se enlaza desde las dos páginas.

La gracia de esto es que no hay nada escrito a mano: los alojamientos, la
agenda, las cifras y las fechas salen del mismo archivo que alimenta la web.
Cuando sustituyas los datos de ejemplo por los reales, vuelve a ejecutarlo y
los catorce archivos se rehacen solos.

Qué genera
----------
  Para todo el mundo        pasaporte-km0.pdf · programa-actividades.pdf
                            bases-sorteo.pdf
  Para alojamientos         carteleria.pdf · kit-redes-sociales.zip
                            textos-para-tu-web.docx · manual-de-marca.pdf
                            sello-pasaporte.pdf · guia-recepcion.pdf
  Para medios               dossier-prensa.pdf · nota-prensa-presentacion.pdf
                            logotipos-y-marca.zip · banco-imagenes.zip
                            alojamientos-adheridos.xlsx

Requisitos: node con playwright (para imprimir el HTML a PDF con las
tipografías de la marca) y openpyxl. Los PDF llevan texto real
seleccionable, no son imágenes.
"""

import os, sys, io, json, zipfile, subprocess, datetime

RAIZ = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
TMP = os.path.join(RAIZ, "_build", "_desc")
HTML = os.path.join(TMP, "html")
SALIDA = os.path.join(RAIZ, "descargas")
FUENTES = os.path.join(RAIZ, "assets", "fonts")

for d in (TMP, HTML, SALIDA):
    os.makedirs(d, exist_ok=True)


# ---------------------------------------------------------------------------
# Los datos, leídos de contenido/*.json — los mismos que alimentan la web,
# así que las cifras de los PDF siempre cuadran con lo que se ve publicado.
# ---------------------------------------------------------------------------
import contenido

ALOJ = contenido.ALOJAMIENTOS
AGENDA = contenido.AGENDA
CFG = contenido.CONFIG

FECHAS = contenido.FECHAS_ES
EMAIL = contenido.EMAIL
TEL = contenido.TELEFONO
WEB = contenido.DOMINIO.split("//", 1)[-1]
HOSBEC = "Asociación Empresarial Hotelera y Turística de la Comunidad Valenciana"

CUPO_TOTAL = contenido.CUPO_TOTAL
DESTINOS = contenido.DESTINOS
PROVINCIAS = ["Castelló", "València", "Alicante"]
HOY = datetime.date(2026, 8, 18)

DIAS_SEM = ["lunes", "martes", "miércoles", "jueves", "viernes", "sábado", "domingo"]
MESES = ["enero", "febrero", "marzo", "abril", "mayo", "junio", "julio",
         "agosto", "septiembre", "octubre", "noviembre", "diciembre"]

INICIO = datetime.date(*[int(x) for x in CFG["fechaInicio"][:10].split("-")])
FIN = datetime.date(*[int(x) for x in CFG["fechaFin"][:10].split("-")])
TOTAL_DIAS = (FIN - INICIO).days + 1


def fecha_dia(n):
    return INICIO + datetime.timedelta(days=n - 1)


def dia_largo(n):
    f = fecha_dia(n)
    return "%s %d de %s" % (DIAS_SEM[f.weekday()], f.day, MESES[f.month - 1])


def es_finde(n):
    return fecha_dia(n).weekday() >= 4      # viernes, sábado, domingo


def tipo_es(t):
    return {"hotel": "Hotel", "apartamentos": "Apartamentos", "camping": "Camping",
            "rural": "Casa rural", "hostal": "Hostal", "balneario": "Balneario"}.get(t, t)


def exp_es(e):
    return {"gastronomia": "Gastronomía", "bienestar": "Bienestar", "familia": "En familia",
            "cultura": "Cultura y pueblos", "mar": "Junto al mar", "deporte": "Naturaleza y deporte",
            "romantico": "En pareja", "mascotas": "Con mascota", "accesible": "Accesible",
            "sostenible": "Sostenible", "noche": "De noche"}.get(e, e)


def L(v, idioma="es"):
    if isinstance(v, dict):
        return v.get(idioma, v.get("es", ""))
    return v or ""


def esc(v):
    return (str(v).replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;"))


# ---------------------------------------------------------------------------
# La piel de todos los documentos: la misma paleta y las mismas tipografías
# que la web, incrustadas para que el PDF sea autosuficiente.
# ---------------------------------------------------------------------------
CSS = """
@font-face { font-family:"Montserrat"; font-weight:300 900;
  src:url("FUENTES/montserrat-latin-wght-normal.woff2") format("woff2"); }
@font-face { font-family:"Montserrat"; font-weight:300 900;
  src:url("FUENTES/montserrat-latin-ext-wght-normal.woff2") format("woff2");
  unicode-range:U+0100-02BA,U+1E00-1E9F,U+2C60-2C7F; }
@font-face { font-family:"Lora"; font-weight:400 700;
  src:url("FUENTES/lora-latin-wght-normal.woff2") format("woff2"); }
@font-face { font-family:"Lora"; font-weight:400 700; font-style:italic;
  src:url("FUENTES/lora-latin-wght-italic.woff2") format("woff2"); }
@font-face { font-family:"Caveat Brush"; font-weight:400;
  src:url("FUENTES/caveat-brush-latin-400-normal.woff2") format("woff2"); }

:root{
  --mar:#1EA4C6; --mar-d:#14647D; --mar-p:#E2F3F8;
  --verde:#8CB26F; --verde-d:#6E9553; --verde-p:#EDF3E4;
  --arena:#EED8AE; --arena-p:#FBF2DE;
  --terra:#D9794D; --terra-h:#C4653A; --terra-p:#FBE7DB;
  --roto:#FAF5EC; --tinta:#123C4C; --texto:#4C6672; --suave:#7D8F98;
  --linea:#E4E9EB;
}
*{box-sizing:border-box;margin:0;}
html,body{-webkit-print-color-adjust:exact; print-color-adjust:exact;}
body{font-family:"Montserrat",sans-serif; color:var(--texto); font-size:10.5pt; line-height:1.55;}
h1,h2,h3,h4{font-family:"Lora",Georgia,serif; color:var(--tinta); line-height:1.15; font-weight:600;}
h1{font-size:30pt;} h2{font-size:17pt;} h3{font-size:12.5pt;} h4{font-size:11pt;}
strong{color:var(--tinta);}
a{color:var(--mar-d); text-decoration:none;}
em{font-family:"Lora",serif;}

/* Columna flexible: el pie se va al fondo con margin-top:auto en vez de ir
   posicionado en absoluto, que era lo que hacia que el texto se le montara
   encima cuando una pagina se pasaba de largo. */
.hoja{page-break-after:always; display:flex; flex-direction:column; min-height:256mm;}
.hoja:last-child{page-break-after:auto;}
/* «suelta» = documento que fluye en varias páginas y lleva el pie repetido
   en el margen, puesto por Chromium, en vez de uno propio al final. */
.hoja.suelta{min-height:0; display:block;}
/* Sin esto, los hijos del flex se encogen por debajo de su contenido: la caja
   mide 264 mm pero por dentro rebosa y al imprimir se va a otra página. */
.hoja > *{flex:0 0 auto;}
/* Para documentos densos que tienen que caber en una sola página */
.hoja.compacta{font-size:9.4pt;}
.hoja.compacta h1{font-size:20pt;}
.hoja.compacta h2{font-size:13.5pt; margin-top:.85em; margin-bottom:.3em;}
.hoja.compacta .lede{font-size:10.4pt;}
.hoja.compacta p + p{margin-top:.5em;}

/* --- marca --- */
.logo{font-family:"Montserrat",sans-serif; font-weight:800; font-size:15pt;
  color:var(--tinta); letter-spacing:-.01em; display:inline-flex; align-items:baseline; gap:.18em;}
.logo .w{font-family:"Caveat Brush",cursive; font-weight:400; color:var(--mar); font-size:16pt;}
.logo.blanco, .logo.blanco .w{color:#fff;}
.logo.grande{font-size:30pt;} .logo.grande .w{font-size:33pt;}
.logo.enorme{font-size:58pt;} .logo.enorme .w{font-size:64pt;}

.ante{display:block; font-size:7.6pt; font-weight:800; letter-spacing:.16em;
  text-transform:uppercase; color:var(--verde-d);}
.ante.mar{color:var(--mar);} .ante.terra{color:var(--terra);} .ante.arena{color:#B58A3C;}
.hand{font-family:"Caveat Brush",cursive; color:var(--terra); font-size:14pt;}

.cab{display:flex; justify-content:space-between; align-items:flex-start;
  border-bottom:2.5px solid var(--tinta); padding-bottom:9px; margin-bottom:20px;}
.cab .fechas{font-size:8.5pt; color:var(--suave); text-align:right; line-height:1.4;}
.pie{margin-top:auto; padding-top:7px; border-top:1px solid var(--linea);
  font-size:7.6pt; color:var(--suave); display:flex; justify-content:space-between;
  flex:0 0 auto;}

.lede{font-size:12pt; line-height:1.6; color:var(--tinta);}
p + p{margin-top:.62em;}
h2{margin-top:1.5em; margin-bottom:.5em;}
h3{margin-top:1.2em; margin-bottom:.35em;}
.hoja > h2:first-child, .hoja > *:first-child{margin-top:0;}

ul,ol{padding-left:1.15em; margin-top:.5em;}
li{margin-bottom:.34em;}
ul.limpia{list-style:none; padding-left:0;}
ul.limpia li{padding-left:1.1em; position:relative;}
ul.limpia li::before{content:""; position:absolute; left:0; top:.52em;
  width:6px; height:6px; border-radius:50%; background:var(--mar);}

.cols2{column-count:2; column-gap:9mm;}
.rejilla{display:grid; gap:5mm;}
.r2{grid-template-columns:1fr 1fr;} .r3{grid-template-columns:1fr 1fr 1fr;}
.r4{grid-template-columns:repeat(4,1fr);}

.rejilla{flex:0 0 auto;}
.caja{border:1.5px solid var(--linea); border-radius:5mm; padding:5mm; break-inside:avoid;}
.caja.arena{background:var(--arena-p); border-color:#E7D6B4;}
.caja.mar{background:var(--mar-p); border-color:#CDE6F0;}
.caja.verde{background:var(--verde-p); border-color:#DCE8CF;}
.caja.tinta{background:var(--tinta); border-color:var(--tinta); color:#EAF6F9;}
.caja.tinta h2,.caja.tinta h3{color:#fff;}

.cifra{font-family:"Lora",serif; font-size:26pt; font-weight:600; color:var(--mar); line-height:1;}
.cifra + .u{display:block; font-size:8.6pt; font-weight:700; color:var(--tinta); margin-top:3px;}
.u + .n{display:block; font-size:7.8pt; color:var(--suave); margin-top:1px;}

table{width:100%; border-collapse:collapse; font-size:9pt;}
th{text-align:left; background:var(--tinta); color:#fff; font-size:7.6pt;
  letter-spacing:.09em; text-transform:uppercase; padding:6px 7px; font-weight:700;}
td{padding:5.5px 7px; border-bottom:1px solid var(--linea); vertical-align:top;}
tr:nth-child(even) td{background:#FBFCFC;}

.aviso{border-left:4px solid var(--terra); background:var(--terra-p);
  padding:4mm 5mm; border-radius:0 3mm 3mm 0; font-size:9pt; margin:5mm 0;}
.aviso b{color:var(--terra-h);}

.dia{break-inside:avoid; margin-bottom:4mm;}
.dia-cab{display:flex; align-items:baseline; gap:8px; border-bottom:1.5px solid var(--arena);
  padding-bottom:3px; margin-bottom:3mm;}
.dia-cab .n{font-family:"Lora",serif; font-size:14pt; font-weight:600; color:var(--tinta);}
.dia-cab .s{font-size:8pt; color:var(--suave); text-transform:uppercase; letter-spacing:.1em;}
.dia-cab .f{margin-left:auto; font-size:7.5pt; color:#B58A3C; font-weight:700;
  text-transform:uppercase; letter-spacing:.1em;}
.acto{display:grid; grid-template-columns:16mm 1fr 26mm; gap:4mm; padding:2.5mm 0;
  border-bottom:1px solid var(--linea); break-inside:avoid;}
.acto .h{font-family:"Lora",serif; font-weight:600; color:var(--mar-d);}
.acto .t{font-weight:700; color:var(--tinta);}
.acto .d{font-size:8.6pt;}
.acto .p{text-align:right; font-size:8.6pt; font-weight:700; color:var(--terra-h);}
.acto .p.gratis{color:var(--verde-d);}

.muestra{height:22mm; border-radius:3mm; margin-bottom:2.5mm;}
.hex{font-family:ui-monospace,Menlo,Consolas,monospace; font-size:8.5pt; color:var(--suave);}
"""


def envolver(titulo, cuerpo, css_extra="", tam="A4"):
    """Documento HTML completo listo para imprimir."""
    return """<!DOCTYPE html><html lang="es"><head><meta charset="utf-8">
<title>%s</title><style>%s
@page{size:%s; margin:14mm 15mm 13mm;}
%s</style></head><body>%s</body></html>""" % (
        esc(titulo), CSS.replace("FUENTES", "file://" + FUENTES), tam, css_extra, cuerpo)


def cabecera(titulo_corto, ante=""):
    return """<div class="cab">
  <div><span class="logo"><span>KM0</span><span class="w">week</span></span>
    %s</div>
  <div class="fechas"><b style="color:#123C4C">%s</b><br>Comunitat Valenciana</div>
</div>""" % ('<div class="ante" style="margin-top:2px">%s</div>' % esc(ante) if ante else "",
             esc(FECHAS))


def pie(txt=""):
    return """<div class="pie"><span>%s</span><span>HOSBEC · %s · %s</span></div>""" % (
        esc(txt or "HOSBEC Km0 Week " + FECHAS), esc(EMAIL), esc(TEL))


# ===========================================================================
# 1 · PROGRAMA DE ACTIVIDADES
# ===========================================================================
def doc_programa():
    dias = sorted({a["dia"] for a in AGENDA})
    bloques = []
    for n in dias:
        actos = sorted([a for a in AGENDA if a["dia"] == n], key=lambda a: a["hora"])
        filas = "".join(
            """<div class="acto">
                 <div class="h">%s</div>
                 <div><div class="t">%s</div><div class="d">%s</div>
                      <div class="d" style="color:#7D8F98">%s</div></div>
                 <div class="p%s">%s</div>
               </div>""" % (
                esc(a["hora"]), esc(L(a["titulo"])), esc(L(a["desc"])), esc(a["lugar"]),
                " gratis" if L(a["precio"]).lower().startswith("gratis") else "",
                esc(L(a["precio"])))
            for a in actos)
        f = fecha_dia(n)
        bloques.append(
            """<div class="dia"><div class="dia-cab">
                 <span class="n">%d</span>
                 <span class="s">%s · %s</span>
                 %s</div>%s</div>""" % (
                f.day, DIAS_SEM[f.weekday()], MESES[f.month - 1],
                '<span class="f">fin de semana</span>' if es_finde(n) else "", filas))

    gratis = sum(1 for a in AGENDA if L(a["precio"]).lower().startswith("gratis"))
    cuerpo = """<div class="hoja suelta">%s
  <span class="ante">Programa de actividades</span>
  <h1 style="margin:3px 0 10px">Diecisiete días,<br>algo que hacer cada fin de semana</h1>
  <p class="lede" style="max-width:150mm">Las %d actividades abiertas de la primera Km0 Week.
     No hace falta alojarse para venir a ninguna: son abiertas a todo el mundo.
     %d son gratuitas y el resto tienen una aportación a coste.</p>

  <div class="rejilla r4" style="margin:7mm 0 6mm">
    <div class="caja arena"><span class="cifra" style="color:#B58A3C">%d</span>
      <span class="u">actividades</span><span class="n">en toda la Comunitat</span></div>
    <div class="caja verde"><span class="cifra" style="color:#6E9553">%d</span>
      <span class="u">gratuitas</span><span class="n">sin coste ninguno</span></div>
    <div class="caja mar"><span class="cifra">%d</span>
      <span class="u">días de edición</span><span class="n">tres fines de semana</span></div>
    <div class="caja"><span class="cifra" style="color:#D9794D">%d</span>
      <span class="u">municipios</span><span class="n">costa e interior</span></div>
  </div>

  <div class="aviso"><b>Antes de venir.</b> Casi todas requieren reserva previa y las plazas
    son limitadas. El enlace de inscripción de cada actividad se abre quince días antes del
    arranque en %s. Si llueve mucho, algunas se trasladan o se cancelan: te avisamos por el
    mismo canal donde te hayas inscrito.</div>
  %s
  <div class="caja tinta" style="margin-top:7mm">
    <h3 style="margin:0 0 3mm">¿Organizas algo y quieres que entre en el programa?</h3>
    <p style="color:#EAF6F9">Si eres ayuntamiento, alojamiento adherido o entidad de la
      Comunitat y tienes una propuesta abierta al público, escríbenos a %s. Entra en el
      programa si es abierta, si es gratuita o a coste, y si sucede entre el %s.</p>
  </div>
  %s
</div>""" % (
        cabecera("Programa", "Actividades abiertas"),
        len(AGENDA), gratis, len(AGENDA), gratis, TOTAL_DIAS,
        len({a["lugar"] for a in AGENDA}),
        WEB,
        "".join(bloques),
        esc(EMAIL), esc(FECHAS), "")
    return envolver("Programa de actividades · HOSBEC Km0 Week", cuerpo)


# ===========================================================================
# 2 · PASAPORTE IMPRIMIBLE (A4 apaisado, se dobla por la mitad)
# ===========================================================================
def doc_pasaporte():
    casillas = "".join(
        """<div style="border:1.6px dashed #C9D6DB; border-radius:4mm; height:56mm;
             display:flex; flex-direction:column; align-items:center; justify-content:center;
             background:#fff">
             <span style="font-family:'Lora',serif; font-size:22pt; color:#C9D6DB">%d</span>
             <span style="font-size:7pt; color:#B6C4CB; letter-spacing:.08em">SELLO</span>
           </div>""" % i for i in range(1, 8))
    # La octava celda no es un sello: cierra la rejilla y recoge lo que el
    # visitante quiera apuntarse (qué visitó, con quién, qué le gustó).
    casillas += """<div style="border:1.6px solid #EED8AE; border-radius:4mm; height:56mm;
         background:#FDF7EC; padding:5mm 6mm; display:flex; flex-direction:column">
         <b style="font-size:8.6pt; color:#123C4C">Tus notas</b>
         <span style="font-size:7.2pt; color:#9C8A63; margin-bottom:3mm">Qué visitaste,
           con quién, qué repetirías.</span>
         %s
       </div>""" % ('<div style="border-bottom:1px solid #E2D3B0; height:6.5mm"></div>' * 4)

    cuerpo = """
<div class="hoja" style="display:grid; grid-template-columns:1fr 1fr; gap:0;
     height:210mm; min-height:0">

  <!-- contraportada (queda detrás al doblar) -->
  <div style="padding:11mm 12mm; border-right:1px dashed #C9D6DB; display:flex;
       flex-direction:column; justify-content:center">
    <h3 style="margin:0 0 3mm">Cómo funciona</h3>
    <ul class="limpia" style="font-size:9pt">
      <li>Cada estancia en un alojamiento adherido suma <b>un sello</b>.</li>
      <li>Cada actividad del programa suma <b>un sello</b>.</li>
      <li>Con <b>tres sellos</b> entras en el sorteo de diez estancias de fin de semana
          para dos personas, para disfrutar en 2027.</li>
      <li>Con <b>cinco sellos</b> entras además en el sorteo de la cena de clausura.</li>
      <li>Se cierra el <b>domingo 29 de noviembre a las 14:00</b>. El sorteo se celebra
          ese mismo día a las 18:00, ante notario y retransmitido.</li>
    </ul>

    <h3>Tus datos</h3>
    <p style="font-size:8.4pt; color:#7D8F98; margin-bottom:3mm">Rellénalos para poder
      avisarte si resultas premiado.</p>
    %s

    <p style="font-size:7.2pt; color:#7D8F98; margin-top:5mm; line-height:1.45">
      Los datos se usan únicamente para gestionar el sorteo y se destruyen después.
      Responsable: HOSBEC, %s. Bases completas en %s/descargas.html</p>
  </div>

  <!-- portada -->
  <div style="padding:13mm 12mm; background:#123C4C; color:#EAF6F9; display:flex;
              flex-direction:column; justify-content:space-between">
    <div>
      <span class="ante arena">Pasaporte</span>
      <div class="logo grande blanco" style="margin:2mm 0 4mm"><span>KM0</span><span class="w"
        style="color:#EED8AE">week</span></div>
      <p style="font-size:10pt; color:#EAF6F9; max-width:70mm">Un cuaderno de bolsillo para
        sellar lo que vas descubriendo cerca de casa.</p>
      <p class="hand" style="color:#EED8AE; font-size:13pt; margin-top:4mm">
        Descubre lo cerca, vive lo nuestro.</p>
    </div>
    <div style="font-size:8pt; color:rgba(234,246,249,.72); line-height:1.5">
      <b style="color:#EED8AE">%s</b><br>
      Comunitat Valenciana<br>%s
    </div>
  </div>
</div>

<div class="hoja" style="height:210mm; min-height:0; padding:14mm 16mm;
     justify-content:center">
  <div style="text-align:center; margin-bottom:5mm">
    <span class="ante mar">Siete sellos</span>
    <h2 style="margin:1mm 0 0">Cada plan, un sello</h2>
  </div>
  <div class="rejilla" style="grid-template-columns:repeat(4,1fr); gap:5mm">%s</div>
  <p style="text-align:center; font-size:8pt; color:#7D8F98; margin-top:5mm">
    Pide el sello en la recepción del alojamiento o a quien guíe la actividad.
    Imprime en A4, dobla por la mitad y llévalo encima.</p>
</div>""" % (
        "".join(
            '<div style="border-bottom:1px solid #C9D6DB; height:7mm; margin-bottom:3mm;'
            ' font-size:7.4pt; color:#7D8F98">%s</div>' % c
            for c in ("Nombre y apellidos", "Correo electrónico", "Teléfono", "Municipio")),
        esc(HOSBEC), WEB, esc(FECHAS), esc(EMAIL), casillas)
    return envolver("Pasaporte Km0 · HOSBEC Km0 Week", cuerpo,
                    css_extra="@page{size:A4 landscape; margin:0;}")


# ===========================================================================
# 3 · BASES DEL SORTEO
# ===========================================================================
def doc_bases():
    cuerpo = """<div class="hoja suelta">%s
  <span class="ante terra">Bases legales</span>
  <h1 style="margin:3px 0 8px; font-size:26pt">Bases del sorteo<br>del Pasaporte Km0</h1>
  <p class="lede">Sorteo de diez estancias de fin de semana y de una cena de clausura,
     asociado al Pasaporte Km0 de la HOSBEC Km0 Week %s.</p>

  <div class="aviso"><b>Documento pendiente de validación jurídica.</b> Este texto es un
    borrador completo, redactado para que el asesor jurídico de HOSBEC lo revise y lo cierre.
    No debe publicarse ni distribuirse sin ese visto bueno: un sorteo promocional tiene
    obligaciones concretas en materia de consumo, protección de datos y fiscalidad, y algunos
    apartados (marcados con corchetes) requieren datos que solo HOSBEC puede fijar.</div>

  <h2>1 · Entidad organizadora</h2>
  <p>HOSBEC, %s, con CIF G03270014 y domicilio en Paseo Els Tolls, 2 (Edificio INVATTUR,
     3.ª planta), 03502 Benidorm (Alicante), organiza el presente sorteo con finalidad
     promocional, sin coste de participación ni obligación de compra más allá de la propia
     participación en las actividades de la iniciativa.</p>

  <h2>2 · Ámbito y duración</h2>
  <p>El sorteo se desarrolla en el ámbito territorial de la Comunitat Valenciana. El periodo
     de participación comprende del <b>%s</b>, ambos incluidos. La recogida de sellos se
     cierra el <b>domingo 29 de noviembre de 2026 a las 14:00</b>.</p>

  <h2>3 · Quién puede participar</h2>
  <ul class="limpia">
    <li>Personas físicas mayores de 18 años.</li>
    <li>Con residencia acreditable en la Comunitat Valenciana.</li>
    <li>Que hayan completado un Pasaporte Km0 con el número de sellos exigido.</li>
  </ul>
  <p>Quedan excluidos el personal de HOSBEC, el de los alojamientos adheridos y el de las
     entidades colaboradoras, así como sus familiares hasta el segundo grado.</p>

  <h2>4 · Mecánica</h2>
  <p>Cada estancia en un alojamiento adherido y cada asistencia a una actividad del programa
     dan derecho a un sello en el Pasaporte Km0, que se solicita en la recepción del
     alojamiento o a la persona responsable de la actividad. Un mismo establecimiento o
     actividad no puede sellar dos veces el mismo pasaporte.</p>
  <ul class="limpia">
    <li><b>Tres sellos</b> dan acceso al sorteo de las diez estancias.</li>
    <li><b>Cinco sellos</b> dan acceso además al sorteo de la cena de clausura.</li>
  </ul>
  <p>El pasaporte cumplimentado se entrega en cualquier alojamiento adherido antes del cierre,
     o se remite escaneado a %s dentro del mismo plazo.</p>

  <h2>5 · Premios</h2>
  <table style="margin-top:3mm">
    <tr><th>Premio</th><th>Cantidad</th><th>Descripción</th><th>Valor unitario</th></tr>
    <tr><td>Estancia de fin de semana</td><td>10</td>
        <td>Dos noches para dos personas en un alojamiento adherido, con desayuno,
            a disfrutar durante 2027 según disponibilidad</td><td>[por determinar]</td></tr>
    <tr><td>Cena de clausura</td><td>[nº plazas]</td>
        <td>Invitación doble a la cena de clausura de la edición</td>
        <td>[por determinar]</td></tr>
  </table>
  <p style="margin-top:3mm">Los premios son personales e intransferibles y no son canjeables
     por su valor en metálico. Las fechas concretas de disfrute se acuerdan entre la persona
     premiada y el alojamiento, sujetas a disponibilidad y excluyendo los periodos de alta
     ocupación que cada establecimiento comunique previamente a HOSBEC.</p>
  <h2>6 · Sorteo y comunicación</h2>
  <p>El sorteo se celebra el <b>domingo 29 de noviembre de 2026 a las 18:00</b>, ante notario
     y con retransmisión pública, mediante selección aleatoria entre todos los pasaportes
     válidamente presentados. Se extraen diez ganadores y [nº] suplentes por orden.</p>
  <p>Las personas premiadas serán avisadas en las 72 horas siguientes por correo electrónico y
     teléfono, con los datos facilitados en el pasaporte. Disponen de <b>diez días naturales</b>
     para aceptar el premio; transcurrido ese plazo sin respuesta, el premio pasa al primer
     suplente disponible.</p>

  <h2>7 · Protección de datos</h2>
  <p>Los datos personales recogidos en el pasaporte (nombre y apellidos, correo electrónico,
     teléfono y municipio) son tratados por HOSBEC como responsable, con la única finalidad de
     gestionar la participación en el sorteo, comunicar el resultado y entregar los premios.
     La base jurídica es el consentimiento que se presta al entregar el pasaporte.</p>
  <p>Los datos se conservan hasta la entrega de los premios y, como máximo, seis meses después
     del sorteo, tras lo cual se suprimen. No se ceden a terceros salvo al alojamiento donde
     se disfrute el premio y en la medida imprescindible para prestarlo. Puedes ejercer los
     derechos de acceso, rectificación, supresión, limitación, oposición y portabilidad
     escribiendo a %s, y reclamar ante la Agencia Española de Protección de Datos.</p>

  <h2>8 · Fiscalidad</h2>
  <p>Los premios de esta naturaleza pueden estar sujetos a la normativa fiscal aplicable a
     premios y ganancias patrimoniales. [Pendiente de que el asesor fiscal de HOSBEC concrete
     el tratamiento aplicable en función del valor final de los premios, la obligación o no de
     practicar ingreso a cuenta y la información que deba facilitarse a las personas
     premiadas.]</p>

  <h2>9 · Derechos de imagen</h2>
  <p>La aceptación del premio conlleva la autorización, siempre revocable, a que HOSBEC difunda
     el nombre de la persona premiada en los canales de la iniciativa. El uso de imágenes
     requerirá autorización expresa e independiente.</p>

  <h2>10 · Reserva y aceptación</h2>
  <p>HOSBEC se reserva el derecho a modificar estas bases por causa justificada, dando la misma
     publicidad que a las originales, y a excluir a quien incumpla las condiciones o actúe de
     forma fraudulenta. La participación implica la aceptación íntegra de estas bases.</p>

  <h2>11 · Legislación y fuero</h2>
  <p>Estas bases se rigen por la legislación española. Para cualquier controversia, las partes
     se someten a los juzgados y tribunales que correspondan conforme a derecho.</p>

  <p style="margin-top:8mm; font-size:8.4pt; color:#7D8F98">
    Bases depositadas ante [notaría] con fecha [fecha]. Texto completo disponible en
    %s/descargas.html · Consultas en %s</p>
  %s
</div>""" % (
        cabecera("Bases", "Sorteo Pasaporte Km0"), CFG["edicion"],
        esc(HOSBEC), esc(FECHAS), esc(EMAIL),
        esc(EMAIL), WEB, esc(EMAIL), "")
    return envolver("Bases del sorteo · HOSBEC Km0 Week", cuerpo)


# ===========================================================================
# 4 · GUÍA RÁPIDA PARA RECEPCIÓN (una hoja)
# ===========================================================================
def doc_guia():
    cuerpo = """<div class="hoja">%s
  <span class="ante mar">Para tu equipo de recepción</span>
  <h1 style="margin:3px 0 6px; font-size:25pt">Km0 Week en una hoja</h1>
  <p class="lede">Imprímela y déjala en el mostrador. Es todo lo que hace falta saber.</p>

  <div class="rejilla r2" style="margin-top:6mm">
    <div class="caja mar">
      <h3 style="margin-top:0">¿Qué es?</h3>
      <p style="font-size:9.5pt">Del <b>%s</b>, los alojamientos de la Comunitat abrimos las
        puertas a nuestros propios vecinos con una oferta pensada para ellos. Tres fines de
        semana. Lo organiza HOSBEC y participan %d alojamientos de las tres provincias.</p>
    </div>
    <div class="caja verde">
      <h3 style="margin-top:0">¿Quién puede reservar?</h3>
      <p style="font-size:9.5pt">Cualquier persona <b>residente en la Comunitat Valenciana</b>.
        Se acredita con el DNI o cualquier documento con domicilio en la Comunitat, en el
        momento del check-in. No hace falta nada más: ni códigos, ni registros previos.</p>
    </div>
  </div>

  <h2>Las cinco cosas que nos preguntan</h2>
  <div class="rejilla r2" style="gap:4mm">
    <div><h4>«¿El precio es de verdad más barato?»</h4>
      <p style="font-size:9.3pt">Sí. El compromiso es que el precio Km0 sea igual o inferior
        al más bajo que ha tenido el alojamiento en los tres meses anteriores para el mismo
        día de la semana y el mismo tipo de habitación. HOSBEC lo comprueba con muestras
        aleatorias durante la edición.</p></div>
    <div><h4>«¿Puedo reservar por una agencia?»</h4>
      <p style="font-size:9.3pt">No. La oferta Km0 es <b>solo reserva directa</b>, por teléfono
        o por nuestra web. Ni comisiones ni intermediarios: por eso el precio puede ser el que
        es.</p></div>
    <div><h4>«¿Cuántas habitaciones hay?»</h4>
      <p style="font-size:9.3pt">Las que figuran como cupo en nuestra ficha de la web, ni una
        menos. Cuando se agotan, se agotan: es un compromiso público, no una estimación.</p></div>
    <div><h4>«¿Qué es el pasaporte?»</h4>
      <p style="font-size:9.3pt">Un cuadernillo con siete casillas. Cada estancia y cada
        actividad suman un sello. Con tres entran en el sorteo de diez estancias para 2027.
        <b>Sellamos nosotros, en recepción.</b></p></div>
  </div>

  <div class="caja arena" style="margin-top:5mm">
    <h3 style="margin-top:0">Cómo se sella</h3>
    <ol style="font-size:9.5pt; margin-top:2mm">
      <li>El cliente enseña el pasaporte al hacer el check-out (o cuando lo pida).</li>
      <li>Sellamos <b>una sola casilla</b>, la primera que esté libre. Un pasaporte no se
          sella dos veces en el mismo alojamiento.</li>
      <li>Si no lleva pasaporte, tenemos ejemplares impresos en el mostrador.</li>
    </ol>
  </div>

  <div class="caja tinta" style="margin-top:5mm">
    <h3 style="margin:0 0 2mm">Si surge cualquier duda</h3>
    <p style="color:#EAF6F9; font-size:9.5pt">Escribe a <b style="color:#EED8AE">%s</b> o llama
      al <b style="color:#EED8AE">%s</b>. Toda la información pública está en %s</p>
  </div>
  %s
</div>""" % (
        cabecera("Guía rápida", "Guía rápida"),
        esc(FECHAS), len(ALOJ), esc(EMAIL), esc(TEL), WEB,
        pie("Guía rápida para recepción · Km0 Week 2026"))
    return envolver("Guía rápida para recepción · HOSBEC Km0 Week", cuerpo)


# ===========================================================================
# 5 · SELLO PARA EL PASAPORTE (plantilla para el fabricante)
# ===========================================================================
def doc_sello():
    cuerpo = """<div class="hoja">%s
  <span class="ante terra">Plantilla de encargo</span>
  <h1 style="margin:3px 0 6px; font-size:25pt">El sello de tu casa</h1>
  <p class="lede" style="max-width:145mm">Lleva esta hoja a cualquier tienda de sellos de goma.
     Con las medidas y el contenido de aquí, te lo hacen en un día.</p>

  <div class="rejilla r2" style="margin-top:7mm; align-items:start">
    <div>
      <h3 style="margin-top:0">Especificaciones</h3>
      <table style="margin-top:3mm">
        <tr><td style="width:38%%"><b>Formato</b></td><td>Redondo</td></tr>
        <tr><td><b>Diámetro</b></td><td>40 mm</td></tr>
        <tr><td><b>Tipo</b></td><td>Automático (entintado incorporado)</td></tr>
        <tr><td><b>Tinta</b></td><td>Azul Mediterráneo · Pantone aproximado 632 C · #1EA4C6</td></tr>
        <tr><td><b>Alternativa</b></td><td>Negro, si la tienda no hace color</td></tr>
        <tr><td><b>Grosor mínimo de línea</b></td><td>0,4 mm</td></tr>
      </table>

      <h3>Qué tiene que poner</h3>
      <ul class="limpia" style="font-size:9.6pt">
        <li><b>Arriba, en el arco:</b> KM0 WEEK</li>
        <li><b>En el centro, grande:</b> el nombre de tu alojamiento</li>
        <li><b>Debajo:</b> tu municipio</li>
        <li><b>Abajo, en el arco:</b> el año de la edición (%s)</li>
      </ul>

      <div class="aviso" style="margin-top:5mm"><b>Un detalle que importa.</b> No metas el
        logotipo de la Km0 Week dentro del sello: a 40 mm no se lee y ensucia la impresión.
        Con el texto «KM0 WEEK» en el arco superior es suficiente y se reconoce igual.</div>
    </div>

    <div style="text-align:center">
      <h3 style="margin-top:0; text-align:left">A tamaño real</h3>
      <div style="display:inline-block; position:relative; margin-top:4mm">
        <div style="width:40mm; height:40mm; border-radius:50%%; border:1.2mm solid #1EA4C6;
                    display:flex; flex-direction:column; align-items:center;
                    justify-content:center; color:#1EA4C6; position:relative">
          <div style="position:absolute; top:3.2mm; font-size:6pt; font-weight:800;
                      letter-spacing:.22em">KM0 WEEK</div>
          <div style="font-family:'Lora',serif; font-size:8.5pt; font-weight:600;
                      line-height:1.1; max-width:28mm">Nombre del<br>alojamiento</div>
          <div style="font-size:6pt; letter-spacing:.12em; margin-top:1mm">MUNICIPIO</div>
          <div style="position:absolute; bottom:3.4mm; font-size:6pt; font-weight:700;
                      letter-spacing:.2em">%s</div>
          <div style="position:absolute; inset:2.4mm; border-radius:50%%;
                      border:.3mm solid #1EA4C6; opacity:.45"></div>
        </div>
        <div style="position:absolute; left:50%%; top:-7mm; transform:translateX(-50%%);
                    font-size:7pt; color:#7D8F98">40 mm</div>
        <div style="position:absolute; left:-3mm; top:-4mm; width:.3mm; height:48mm;
                    background:#C9D6DB"></div>
        <div style="position:absolute; right:-3mm; top:-4mm; width:.3mm; height:48mm;
                    background:#C9D6DB"></div>
      </div>
      <p style="font-size:7.6pt; color:#7D8F98; margin-top:4mm; max-width:70mm;
                margin-inline:auto">Si imprimes esta hoja al 100%% (sin «ajustar a la
        página»), el círculo mide exactamente 40 mm y sirve de referencia.</p>

      <h3 style="text-align:left; margin-top:6mm">Dónde y cómo se usa</h3>
      <ul class="limpia" style="font-size:9.3pt; text-align:left">
        <li><b>En recepción</b>, junto al mostrador. Se sella al hacer el check-out o cuando
            el cliente lo pida.</li>
        <li><b>Una sola casilla</b> por pasaporte y alojamiento: la primera que esté libre.</li>
        <li><b>Que se lea.</b> Presiona en firme sobre superficie dura: un sello borroso da
            problemas si esa persona resulta premiada.</li>
      </ul>
    </div>
  </div>
  %s
</div>""" % (cabecera("Sello", "Sello del pasaporte"), CFG["edicion"], CFG["edicion"],
             pie("Plantilla del sello · Km0 Week 2026"))
    return envolver("Sello para el pasaporte · HOSBEC Km0 Week", cuerpo)


# ===========================================================================
# 6 · CARTELERÍA (A3, A4, faldón y vinilo)
# ===========================================================================
def doc_carteleria():
    def cartel(tam_txt, alto, titular, cuerpo_txt, escala=1.0):
        return """<div class="hoja" style="min-height:%s; display:flex; flex-direction:column;
             justify-content:space-between; background:#123C4C; color:#EAF6F9;
             padding:%smm; margin:0">
          <div>
            <div class="logo blanco" style="font-size:%.1fpt"><span>KM0</span><span class="w"
              style="color:#EED8AE; font-size:%.1fpt">week</span></div>
            <div style="font-size:%.1fpt; letter-spacing:.2em; color:#EED8AE; font-weight:800;
                        margin-top:%.1fmm">%s</div>
          </div>
          <div>
            <h1 style="color:#fff; font-size:%.1fpt; line-height:1.06; max-width:%.0f%%">%s</h1>
            <p style="color:#EAF6F9; font-size:%.1fpt; margin-top:%.1fmm; max-width:%.0f%%">%s</p>
          </div>
          <div style="display:flex; justify-content:space-between; align-items:flex-end;
                      border-top:1px solid rgba(234,246,249,.25); padding-top:%.1fmm">
            <div style="font-size:%.1fpt; line-height:1.5">
              <b style="color:#EED8AE">%s</b><br>%s</div>
            <div style="font-family:'Caveat Brush',cursive; color:#EED8AE; font-size:%.1fpt">
              Descubre lo cerca,<br>vive lo nuestro.</div>
          </div>
        </div>""" % (alto, 16 * escala,
                     20 * escala, 22 * escala, 8 * escala, 5 * escala,
                     esc(tam_txt),
                     46 * escala, 88, titular,
                     13 * escala, 5 * escala, 78, cuerpo_txt,
                     6 * escala, 9 * escala,
                     esc(FECHAS), WEB, 17 * escala)

    faldon = """<div class="hoja" style="min-height:0; height:297mm; margin:0;
         padding:0; justify-content:center">
      <div style="background:#EED8AE; color:#123C4C; padding:12mm 14mm; display:flex;
                  justify-content:space-between; align-items:center; height:99mm">
        <div>
          <div class="logo" style="font-size:26pt"><span>KM0</span><span class="w"
            style="font-size:29pt; color:#C4653A">week</span></div>
          <h2 style="font-size:21pt; margin:4mm 0 2mm; max-width:120mm">Si vives en la
            Comunitat, esta semana la casa es la tuya</h2>
          <p style="font-size:10.5pt; color:#4C6672">Pregúntanos por la tarifa Km0 y por el
            pasaporte. Reserva directa, sin intermediarios.</p>
        </div>
        <div style="text-align:right; font-size:9pt; color:#4C6672; line-height:1.6">
          <b style="color:#123C4C; font-size:11pt">%s</b><br>%s<br>%s
        </div>
      </div>
      <p style="font-size:7.5pt; color:#7D8F98; padding:3mm 14mm">Faldón para mostrador ·
        210 × 99 mm · imprimir en A4 y recortar por el borde inferior de la banda</p>
    </div>""" % (esc(FECHAS), WEB, esc(TEL))

    vinilo = """<div class="hoja" style="min-height:0; height:297mm; margin:0;
         text-align:center; padding:14mm; justify-content:center">
      <p style="font-size:7.5pt; color:#7D8F98; text-align:left; margin-bottom:6mm">
        Vinilo para la puerta · círculo de 178 mm · imprimir en A4 al 100%% y recortar</p>
      <div style="display:inline-block; width:178mm; max-width:100%%; aspect-ratio:1;
           border-radius:50%%; background:#1EA4C6; color:#fff; display:inline-flex;
           flex-direction:column; align-items:center; justify-content:center; padding:18mm">
        <div class="logo blanco" style="font-size:30pt"><span>KM0</span><span class="w"
          style="font-size:33pt; color:#EED8AE">week</span></div>
        <div style="font-family:'Lora',serif; font-size:19pt; font-weight:600; line-height:1.15;
             margin:6mm 0 4mm; max-width:130mm">Aquí somos Km 0</div>
        <p style="font-size:10pt; max-width:120mm; color:#EAF6F9">Alojamiento adherido a la
          Km0 Week · %s</p>
      </div>
    </div>""" % esc(FECHAS)

    cuerpo = (
        cartel("Cartel · A4 · al 141% da un A3", "297mm",
               "Aquí somos<br>Km 0",
               "Tres fines de semana para que quienes vivimos aquí volvamos a dormir en "
               "nuestros hoteles. Pregunta en recepción por la tarifa para residentes.",
               0.92)
        + faldon + vinilo)
    return envolver("Cartelería · HOSBEC Km0 Week", cuerpo,
                    css_extra="@page{size:A4; margin:0;} .hoja{min-height:0; height:297mm}")


# ===========================================================================
# 7 · MANUAL DE MARCA
# ===========================================================================
PALETA = [
    ("Azul bienestar", "#1EA4C6", "Color principal. Titulares de acento, botones, enlaces y el «week» del logotipo."),
    ("Azul tinta", "#123C4C", "Titulares y fondos oscuros. Es el color del texto de marca, no el negro."),
    ("Verde wellness", "#8CB26F", "Naturaleza, sostenibilidad, actividades al aire libre."),
    ("Arena cálida", "#EED8AE", "Fondos secundarios y detalles sobre azul tinta."),
    ("Terracota ocio", "#D9794D", "Llamadas a la acción y acentos escritos a mano. Con moderación."),
    ("Blanco roto", "#FAF5EC", "Fondo de secciones alternas. Nunca blanco puro en bloques grandes."),
]


def doc_manual():
    muestras = "".join(
        """<div style="break-inside:avoid">
             <div class="muestra" style="background:%s; %s"></div>
             <b style="color:#123C4C; font-size:9.6pt">%s</b><br>
             <span class="hex">%s</span>
             <p style="font-size:8.4pt; margin-top:1.5mm">%s</p>
           </div>""" % (hx, "border:1px solid #E4E9EB" if hx in ("#FAF5EC",) else "",
                        esc(n), hx, esc(d))
        for n, hx, d in PALETA)

    cuerpo = """
<div class="hoja" style="background:#123C4C; color:#EAF6F9; margin:-14mm -15mm; padding:34mm 22mm;
     min-height:275mm; display:flex; flex-direction:column; justify-content:space-between">
  <div>
    <span class="ante arena">Manual de marca · Edición %s</span>
    <div class="logo enorme blanco" style="margin:6mm 0 8mm"><span>KM0</span><span class="w"
      style="color:#EED8AE">week</span></div>
    <p style="font-size:14pt; color:#EAF6F9; max-width:120mm; line-height:1.5">Cómo se usa la
      marca de la Km0 Week: el logotipo, los colores, las tipografías y el tono. Cuatro bloques
      de reglas para que todo lo que salga de la iniciativa se reconozca a la primera.</p>
  </div>
  <div style="font-size:9pt; color:rgba(234,246,249,.75); line-height:1.6">
    HOSBEC · %s<br>%s · %s
  </div>
</div>

<div class="hoja">%s
  <h2 style="margin-top:0">1 · El logotipo</h2>
  <p>El logotipo es un <b>lockup de dos piezas</b>: «KM0» en Montserrat ExtraBold y «week» en
     Caveat Brush. Esa mezcla de una tipografía de sistema con una escrita a mano es la idea
     entera de la marca: algo cercano dentro de algo serio.</p>

  <div class="rejilla r2" style="margin-top:5mm">
    <div class="caja"><span class="ante mar">Versión principal</span>
      <div style="text-align:center; padding:8mm 0"><span class="logo grande"><span>KM0</span><span
        class="w">week</span></span></div>
      <p style="font-size:8.6pt">Sobre fondo claro. Es la que se usa siempre que se pueda.</p></div>
    <div class="caja tinta"><span class="ante arena">Versión sobre oscuro</span>
      <div style="text-align:center; padding:8mm 0"><span class="logo grande blanco"><span>KM0</span><span
        class="w" style="color:#EED8AE">week</span></span></div>
      <p style="font-size:8.6pt; color:#EAF6F9">Sobre azul tinta o fotografía oscura.
        El «week» pasa a arena.</p></div>
  </div>

  <h3>Área de respeto</h3>
  <p>Alrededor del logotipo hay que dejar libre, como mínimo, <b>la altura de la «K»</b> por
     los cuatro lados. Nada puede entrar en ese margen: ni texto, ni fotos, ni bordes.</p>

  <h3>Tamaño mínimo</h3>
  <p>En pantalla, <b>110 px</b> de ancho. En impresión, <b>28 mm</b>. Por debajo de eso el
     «week» pierde legibilidad y hay que usar solo el emblema.</p>

  <h3>Lo que no se hace nunca</h3>
  <div class="rejilla r3" style="gap:4mm; margin-top:3mm">
    <div class="caja" style="border-color:#F1C9B8"><b style="color:#C4653A; font-size:9pt">No
      cambiar los colores</b><p style="font-size:8.4pt">El «week» es azul Mediterráneo sobre
      claro y arena sobre oscuro. No hay más versiones.</p></div>
    <div class="caja" style="border-color:#F1C9B8"><b style="color:#C4653A; font-size:9pt">No
      sustituir las tipografías</b><p style="font-size:8.4pt">Ni «week» en otra cursiva, ni
      «KM0» en otra sans. El lockup es una unidad.</p></div>
    <div class="caja" style="border-color:#F1C9B8"><b style="color:#C4653A; font-size:9pt">No
      deformar ni inclinar</b><p style="font-size:8.4pt">Se escala en proporción. Nada de
      sombras, contornos, degradados ni rotaciones.</p></div>
  </div>
  %s
</div>

<div class="hoja">%s
  <h2 style="margin-top:0">2 · La paleta</h2>
  <p>Seis colores. El azul manda, el arena acompaña, la terracota se reserva para lo que hay
     que pulsar. Todos los valores están medidos para cumplir el contraste AA sobre sus fondos
     previstos.</p>
  <div class="rejilla r3" style="margin-top:5mm">%s</div>

  <h3>Proporción de uso</h3>
  <p>Como regla de reparto en cualquier pieza: <b>60%%</b> de blanco o blanco roto,
     <b>25%%</b> de azul (tinta o Mediterráneo), <b>10%%</b> de arena y verde, y no más de un
     <b>5%%</b> de terracota. La terracota que se usa de más deja de llamar la atención, que es
     justo para lo que está.</p>

  <h3>Combinaciones que funcionan</h3>
  <div class="rejilla r4" style="gap:3mm; margin-top:3mm">
    <div style="background:#123C4C; color:#fff; padding:5mm; border-radius:3mm; font-size:8.4pt">
      Texto blanco sobre azul tinta</div>
    <div style="background:#1EA4C6; color:#fff; padding:5mm; border-radius:3mm; font-size:8.4pt">
      Texto blanco sobre Mediterráneo</div>
    <div style="background:#EED8AE; color:#123C4C; padding:5mm; border-radius:3mm; font-size:8.4pt">
      Texto tinta sobre arena</div>
    <div style="background:#FAF5EC; color:#4C6672; padding:5mm; border-radius:3mm;
         font-size:8.4pt; border:1px solid #E4E9EB">Texto cuerpo sobre blanco roto</div>
  </div>
  %s
</div>

<div class="hoja">%s
  <h2 style="margin-top:0">3 · Las tipografías</h2>
  <p>Tres, con un papel muy definido cada una. Las tres son de licencia libre (SIL Open Font
     License), así que se pueden instalar y usar sin coste en cualquier pieza.</p>

  <div class="caja" style="margin-top:5mm">
    <span class="ante mar">Titulares</span>
    <div style="font-family:'Lora',serif; font-size:29pt; color:#123C4C; margin:2mm 0">Lora
      Semibold</div>
    <p style="font-size:9pt">Para titulares y cifras grandes. Es una serif con contraste
      moderado que aporta el punto de oficio que pide la marca. Pesos 400 y 600.</p>
  </div>

  <div class="caja" style="margin-top:4mm">
    <span class="ante mar">Texto e interfaz</span>
    <div style="font-size:24pt; font-weight:600; color:#123C4C; margin:2mm 0">Montserrat
      Medium</div>
    <p style="font-size:9pt">Todo el cuerpo de texto, los rótulos, los botones y las tablas.
      Pesos 500 para cuerpo, 700 para destacados y 800 para antetítulos en versalitas.</p>
  </div>

  <div class="caja" style="margin-top:4mm">
    <span class="ante mar">Acentos</span>
    <div style="font-family:'Caveat Brush',cursive; font-size:29pt; color:#D9794D; margin:2mm 0">
      Caveat Brush</div>
    <p style="font-size:9pt">Solo para el «week» del logotipo y para frases sueltas de tono
      cercano. Nunca en párrafos, nunca en mayúsculas, nunca por debajo de 12 pt.</p>
  </div>

  <h3>Jerarquía tipográfica</h3>
  <table style="margin-top:3mm">
    <tr><th>Uso</th><th>Tipografía</th><th>Tamaño</th><th>Color</th></tr>
    <tr><td>Titular principal</td><td>Lora Semibold</td><td>32–56 pt</td><td>Azul tinta</td></tr>
    <tr><td>Titular de sección</td><td>Lora Semibold</td><td>20–28 pt</td><td>Azul tinta</td></tr>
    <tr><td>Antetítulo</td><td>Montserrat ExtraBold</td><td>7–9 pt · +0,16 em</td><td>Verde o Mediterráneo</td></tr>
    <tr><td>Entradilla</td><td>Montserrat Medium</td><td>12–14 pt</td><td>Azul tinta</td></tr>
    <tr><td>Cuerpo</td><td>Montserrat Medium</td><td>10–11 pt</td><td>Texto</td></tr>
    <tr><td>Pie y notas</td><td>Montserrat Medium</td><td>7,5–9 pt</td><td>Suave</td></tr>
  </table>
  %s
</div>

<div class="hoja">%s
  <h2 style="margin-top:0">4 · El tono</h2>
  <p>La Km0 Week le habla a un vecino, no a un turista. Eso cambia cómo se escribe todo.</p>

  <div class="rejilla r2" style="margin-top:5mm">
    <div class="caja verde"><h4 style="margin-top:0; color:#4A6B32">Sí</h4>
      <ul class="limpia" style="font-size:9.2pt">
        <li>Tutear. Siempre.</li>
        <li>Frases cortas y concretas. «Dormir a veinte minutos de casa».</li>
        <li>Datos comprobables: cupos, precios, distancias.</li>
        <li>Nombrar los sitios por su nombre: Peñíscola, el Palmeral, el Grau.</li>
        <li>Reconocer lo que no se sabe todavía.</li>
      </ul></div>
    <div class="caja" style="border-color:#F1C9B8"><h4 style="margin-top:0; color:#C4653A">No</h4>
      <ul class="limpia" style="font-size:9.2pt">
        <li>«Experiencias únicas», «destinos de ensueño», «vive la magia».</li>
        <li>Exclamaciones y mayúsculas para gritar.</li>
        <li>Descuentos sin decir sobre qué precio.</li>
        <li>Hablar de «producto turístico» a quien va a dormir ahí.</li>
        <li>Anglicismos evitables. Es «reserva directa», no «direct booking».</li>
      </ul></div>
  </div>

  <h3>El claim y el lema</h3>
  <div class="caja arena" style="margin-top:3mm">
    <p style="font-family:'Lora',serif; font-size:17pt; color:#123C4C; line-height:1.25">
      «Todo lo bueno está más cerca de lo que crees»</p>
    <p style="font-size:8.8pt; margin-top:2mm">Claim principal. Se usa en titular, nunca
      partido en dos piezas ni entrecomillado dentro de un texto corrido.</p>
    <p class="hand" style="font-size:16pt; margin-top:4mm">Descubre lo cerca, vive lo nuestro.</p>
    <p style="font-size:8.8pt; margin-top:1mm">Lema de cierre. Va en Caveat Brush y en
      terracota, o en arena sobre fondo oscuro.</p>
  </div>

  <h3>Bilingüismo</h3>
  <p>Todo el material público va en castellano y valencià. En web se resuelve con el selector;
     en pieza impresa, o se hace una versión de cada, o conviven en la misma pieza con el
     valencià en un cuerpo ligeramente menor. Lo que no se hace es traducir solo la mitad.</p>

  <h3>Cómo se escribe el nombre</h3>
  <table style="margin-top:3mm">
    <tr><th>Correcto</th><th>Incorrecto</th><th>Por qué</th></tr>
    <tr><td>Km0 Week</td><td>KM0 WEEK, km0week, Km0week</td><td>En texto corrido, así</td></tr>
    <tr><td>HOSBEC Km0 Week</td><td>Km0 Week by HOSBEC</td><td>La marca madre va delante</td></tr>
    <tr><td>la Km0 Week</td><td>el Km0 Week</td><td>Femenino, por «la semana»</td></tr>
  </table>
  %s
</div>""" % (
        CFG["edicion"], esc(HOSBEC), esc(EMAIL), esc(TEL),
        cabecera("Manual", "Manual de marca"), pie("Manual de marca · Km0 Week 2026"),
        cabecera("Manual", "Manual de marca"), muestras, pie("Manual de marca · Km0 Week 2026"),
        cabecera("Manual", "Manual de marca"), pie("Manual de marca · Km0 Week 2026"),
        cabecera("Manual", "Manual de marca"), pie("Manual de marca · Km0 Week 2026"))
    return envolver("Manual de marca · HOSBEC Km0 Week", cuerpo)


# ===========================================================================
# 8 · DOSSIER DE PRENSA
# ===========================================================================
def doc_dossier():
    por_prov = {p: [a for a in ALOJ if a["provincia"] == p] for p in PROVINCIAS}
    tabla_prov = "".join(
        "<tr><td><b>%s</b></td><td>%d</td><td>%d</td><td>%s</td></tr>" % (
            esc(p), len(v), sum(a.get("cupo") or 0 for a in v),
            esc(", ".join(sorted({a["destino"] for a in v}))))
        for p, v in por_prov.items())

    tipos = {}
    for a in ALOJ:
        tipos[a["tipo"]] = tipos.get(a["tipo"], 0) + 1
    tabla_tipos = "".join(
        "<tr><td>%s</td><td>%d</td></tr>" % (esc(tipo_es(t)), n)
        for t, n in sorted(tipos.items(), key=lambda x: -x[1]))

    cuerpo = """
<div class="hoja" style="background:#123C4C; color:#EAF6F9; margin:-14mm -15mm; padding:32mm 22mm;
     min-height:275mm; display:flex; flex-direction:column; justify-content:space-between">
  <div>
    <span class="ante arena">Dossier de prensa · Edición %s</span>
    <div class="logo grande blanco" style="margin:5mm 0 7mm"><span>KM0</span><span class="w"
      style="color:#EED8AE">week</span></div>
    <h1 style="color:#fff; font-size:34pt; max-width:135mm; line-height:1.1">Todo lo bueno está
      más cerca de lo que crees</h1>
    <p style="font-size:13pt; color:#EAF6F9; max-width:125mm; margin-top:6mm; line-height:1.55">
      Del %s, %d alojamientos de la Comunitat Valenciana abren las puertas a sus propios
      vecinos con %d plazas reservadas y precios que no se encuentran en ningún otro canal.</p>
  </div>
  <div style="font-size:9.5pt; color:rgba(234,246,249,.8); line-height:1.6">
    <b style="color:#EED8AE">Contacto de prensa</b><br>
    HOSBEC · %s<br>%s · %s<br>%s
  </div>
</div>

<div class="hoja">%s
  <h2 style="margin-top:0">Qué es la Km0 Week</h2>
  <p class="lede">Una iniciativa de HOSBEC para que quienes viven en la Comunitat Valenciana
     redescubran los alojamientos de su propio territorio. Diecisiete días de noviembre, tres
     fines de semana enteros, con oferta específica para residentes y reserva directa.</p>

  <p style="margin-top:4mm">La paradoja que la origina es sencilla de enunciar: la Comunitat
     recibe millones de visitantes cada año, y quien vive aquí casi nunca duerme en sus propios
     hoteles. La Km0 Week le da la vuelta a eso durante tres fines de semana de temporada baja,
     cuando hay sitio para recibir bien y tiempo para hablar con quien entra por la puerta.</p>

  <p>Está inspirada en formatos que ya funcionan en otras capitales europeas, como la Madrid
     Hotel Week, con una diferencia de fondo: aquí el protagonista no es una ciudad, sino un
     territorio entero y la gente que vive en él. De Vinaròs a Torrevieja, costa e interior.</p>

  <h2>Las cifras de la primera edición</h2>
  <div class="rejilla r4" style="margin:4mm 0 5mm">
    <div class="caja mar"><span class="cifra">%d</span><span class="u">alojamientos adheridos</span>
      <span class="n">en las tres provincias</span></div>
    <div class="caja verde"><span class="cifra" style="color:#6E9553">%d</span>
      <span class="u">destinos</span><span class="n">costa e interior</span></div>
    <div class="caja arena"><span class="cifra" style="color:#B58A3C">%d</span>
      <span class="u">plazas comprometidas</span><span class="n">cupo publicado y verificable</span></div>
    <div class="caja"><span class="cifra" style="color:#D9794D">%d</span>
      <span class="u">actividades abiertas</span><span class="n">sin necesidad de alojarse</span></div>
  </div>

  <h3>Reparto por provincia</h3>
  <table><tr><th>Provincia</th><th>Alojamientos</th><th>Plazas</th><th>Destinos</th></tr>%s</table>

  <h3>Reparto por tipo de alojamiento</h3>
  <table style="max-width:80mm"><tr><th>Tipo</th><th>Nº</th></tr>%s</table>
  %s
</div>

<div class="hoja">%s
  <h2 style="margin-top:0">Los cinco compromisos</h2>
  <p>Aparecer en la Km0 Week no es un sello decorativo. Cada alojamiento adherido acepta cinco
     reglas, y quien no las cumple queda fuera de la edición siguiente.</p>

  <div class="rejilla r2" style="margin-top:4mm">
    <div class="caja"><h4 style="margin-top:0">1 · El precio no se maquilla</h4>
      <p style="font-size:9.3pt">El precio Km0 es igual o inferior al más bajo que ha tenido el
        alojamiento en los tres meses anteriores para el mismo día de la semana y el mismo tipo
        de habitación. HOSBEC lo comprueba con muestras aleatorias durante la edición.</p></div>
    <div class="caja"><h4 style="margin-top:0">2 · Reserva directa</h4>
      <p style="font-size:9.3pt">Se reserva directamente con el alojamiento, por teléfono o por
        su web. Ni comisiones, ni intermediarios, ni plataformas por medio. Es lo que permite
        que el precio sea el que es.</p></div>
    <div class="caja"><h4 style="margin-top:0">3 · Cupo publicado</h4>
      <p style="font-size:9.3pt">Cada casa compromete un número concreto de plazas para
        residentes y lo publica en su ficha. Si dice treinta, son treinta. Cualquiera puede
        comprobarlo llamando.</p></div>
    <div class="caja"><h4 style="margin-top:0">4 · Condiciones a la vista</h4>
      <p style="font-size:9.3pt">La letra pequeña está entera en la ficha: qué entra, qué no y
        hasta cuándo se puede cancelar. Sin asteriscos.</p></div>
    <div class="caja" style="grid-column:1/-1"><h4 style="margin-top:0">5 · Algo abierto al barrio</h4>
      <p style="font-size:9.3pt">Al menos una experiencia de la casa abierta a todo el mundo,
        aunque no se alojen. De ahí salen las %d actividades del programa: visitas a espacios
        normalmente cerrados, rutas, talleres y jornadas de puertas abiertas.</p></div>
  </div>

  <h2>Por qué en noviembre</h2>
  <p>Noviembre no es temporada alta, y precisamente por eso funciona. Mover viaje de proximidad
     en temporada baja sostiene la ocupación y el empleo cuando más falta hace, sin añadir
     presión al verano. Y da margen para atender bien: no es lo mismo enseñar la cocina de un
     hotel en agosto que en noviembre.</p>

  <h2>Qué no es</h2>
  <p>No es una campaña de descuentos. HOSBEC no comercializa alojamiento, no gestiona reservas
     y no cobra comisión: el precio, la disponibilidad y las condiciones los pone y los mantiene
     cada establecimiento, que es su único responsable. La web es un escaparate con reglas, y
     el botón de cada ficha lleva a la web o al teléfono del alojamiento.</p>

  <div class="aviso"><b>Declaraciones institucionales.</b> Este dossier no incluye citas
    atribuidas. Comunicación de HOSBEC facilita declaraciones de portavoz bajo petición en
    %s.</div>
  %s
</div>

<div class="hoja">%s
  <h2 style="margin-top:0">El Pasaporte Km0</h2>
  <p>Un cuadernillo de bolsillo con siete casillas. Cada estancia en un alojamiento adherido y
     cada actividad del programa suman un sello. Con tres sellos se entra en el sorteo de diez
     estancias de fin de semana para dos personas a disfrutar en 2027; con cinco, además, en el
     de la cena de clausura. Se cierra el domingo 29 de noviembre a las 14:00 y el sorteo se
     celebra ese mismo día a las 18:00, ante notario y retransmitido.</p>

  <h2>Cómo funciona para el visitante</h2>
  <div class="rejilla r3" style="margin-top:4mm">
    <div class="caja mar"><b style="color:#123C4C">1 · Mira qué tienes cerca</b>
      <p style="font-size:9.2pt; margin-top:2mm">La web calcula, en el propio navegador y sin
        enviar la ubicación a ningún servidor, qué alojamientos hay a media hora, una hora y dos
        horas de casa.</p></div>
    <div class="caja mar"><b style="color:#123C4C">2 · Reserva directo</b>
      <p style="font-size:9.2pt; margin-top:2mm">Cada ficha lleva a la web o al teléfono del
        alojamiento. Basta acreditar residencia en la Comunitat al llegar.</p></div>
    <div class="caja mar"><b style="color:#123C4C">3 · Sella el pasaporte</b>
      <p style="font-size:9.2pt; margin-top:2mm">En recepción, al hacer el check-out. Y en cada
        actividad del programa a la que se asista.</p></div>
  </div>

  <h2>Material disponible para medios</h2>
  <ul class="limpia">
    <li><b>Este dossier</b> en PDF y la nota de prensa de presentación.</li>
    <li><b>Listado completo de alojamientos adheridos</b> en hoja de cálculo, con municipio,
        provincia, tipo, cupo comprometido y contacto.</li>
    <li><b>Programa de actividades</b> día a día, con hora, lugar y condiciones de acceso.</li>
    <li><b>Logotipos</b> en vectorial y PNG, con el manual de marca.</li>
    <li><b>Banco de imágenes</b> libre para uso editorial citando la fuente.</li>
  </ul>
  <p>Todo en <b>%s/prensa.html</b></p>

  <div class="caja tinta" style="margin-top:6mm">
    <h3 style="margin:0 0 2mm">Contacto de comunicación</h3>
    <p style="color:#EAF6F9">Para entrevistas, datos concretos, visitas a alojamientos adheridos
      o acompañamiento durante la edición:<br>
      <b style="color:#EED8AE">%s</b> · <b style="color:#EED8AE">%s</b></p>
    <p style="color:rgba(234,246,249,.75); font-size:8.6pt; margin-top:3mm">HOSBEC · %s ·
      CIF G03270014 · Paseo Els Tolls 2, Edificio INVATTUR, 03502 Benidorm (Alicante)</p>
  </div>
  %s
</div>""" % (
        CFG["edicion"], esc(FECHAS), len(ALOJ), CUPO_TOTAL,
        esc(HOSBEC), esc(EMAIL), esc(TEL), WEB,
        cabecera("Dossier", "Dossier de prensa"),
        len(ALOJ), len(DESTINOS), CUPO_TOTAL, len(AGENDA),
        tabla_prov, tabla_tipos, pie("Dossier de prensa · Km0 Week 2026"),
        cabecera("Dossier", "Dossier de prensa"), len(AGENDA), esc(EMAIL),
        pie("Dossier de prensa · Km0 Week 2026"),
        cabecera("Dossier", "Dossier de prensa"), WEB, esc(EMAIL), esc(TEL), esc(HOSBEC),
        pie("Dossier de prensa · Km0 Week 2026"))
    return envolver("Dossier de prensa · HOSBEC Km0 Week", cuerpo)


# ===========================================================================
# 9 · NOTA DE PRENSA
# ===========================================================================
def doc_nota():
    cuerpo = """<div class="hoja compacta">%s
  <div style="display:flex; justify-content:space-between; align-items:baseline">
    <span class="ante terra">Nota de prensa</span>
    <span style="font-size:8.4pt; color:#7D8F98">Benidorm, %d de %s de %d</span>
  </div>
  <h1 style="margin:4px 0 8px; font-size:25pt; max-width:155mm">%d alojamientos de la Comunitat
    reservan %d plazas para que sus vecinos redescubran su propio territorio</h1>
  <p class="lede" style="max-width:155mm">HOSBEC presenta la primera <b>Km0 Week</b>, que se
     celebrará del %s: tres fines de semana con oferta exclusiva para residentes, reserva
     directa y %d actividades abiertas a todo el mundo.</p>

  <hr style="border:0; border-top:1px solid #E4E9EB; margin:4mm 0">

  <p>La Comunitat recibe millones de visitantes cada año y, sin embargo, quien vive aquí casi
     nunca duerme en sus propios hoteles. Con esa paradoja como punto de partida, <b>HOSBEC</b>
     pone en marcha la primera <b>Km0 Week</b>: tres fines de semana de noviembre en los que los
     residentes son los huéspedes.</p>

  <p>Participan <b>%d alojamientos</b> de <b>%d municipios</b> de las tres provincias, que
     comprometen <b>%d plazas</b> para residentes. Ese cupo no es una estimación: cada casa lo
     publica en su ficha y cualquiera puede comprobarlo llamando.</p>

  <h2>Precio verificable y reserva directa</h2>
  <p>El compromiso central es que el precio Km0 sea <b>igual o inferior al más bajo</b> que ha
     tenido ese alojamiento en los tres meses anteriores para el mismo día de la semana y tipo
     de habitación; HOSBEC lo comprueba con muestras aleatorias y quien no cumpla queda fuera
     de la siguiente convocatoria. Todas las reservas son <b>directas con el establecimiento</b>:
     HOSBEC no comercializa alojamiento ni cobra comisión alguna.</p>
  <h2>%d actividades abiertas y un pasaporte</h2>
  <p>Más allá del alojamiento, el programa incluye <b>%d actividades gratuitas o a coste</b>
     abiertas a cualquiera, sin necesidad de pernoctar: visitas a espacios normalmente cerrados,
     rutas guiadas, talleres de artesanía, salidas ornitológicas y puertas abiertas en cocinas
     de hotel. Los participantes pueden además sellar el <b>Pasaporte Km0</b>, que suma un sello
     por cada estancia y cada actividad y da acceso al sorteo de diez estancias para 2027.</p>
  <p>La elección de noviembre no es casual: mover viaje de proximidad fuera de temporada alta
     sostiene la ocupación y el empleo cuando más falta hace, sin añadir presión al verano.</p>

  <div class="caja" style="margin-top:5mm">
    <b style="color:#123C4C">Para el redactor</b>
    <p style="font-size:8.9pt; margin-top:2mm">Este texto se distribuye sin declaraciones
      atribuidas: Comunicación de HOSBEC facilita citas de portavoz, datos desagregados por
      provincia y visitas a alojamientos adheridos bajo petición.</p>
    <p style="font-size:8.9pt; margin-top:2mm">Dossier completo, listado de alojamientos en hoja
      de cálculo, programa, logotipos y banco de imágenes libre para uso editorial en
      <b>%s/prensa.html</b> · <b>Contacto:</b> %s · %s</p>
  </div>
  %s
</div>""" % (
        cabecera("Nota de prensa", ""),
        HOY.day, MESES[HOY.month - 1], HOY.year,
        len(ALOJ), CUPO_TOTAL, esc(FECHAS), len(AGENDA),
        len(ALOJ), len(DESTINOS), CUPO_TOTAL,
        len(AGENDA), len(AGENDA), WEB, esc(EMAIL), esc(TEL),
        pie("Nota de prensa · Km0 Week 2026"))
    return envolver("Nota de prensa · HOSBEC Km0 Week", cuerpo)


# ===========================================================================
# 10 · KIT DE REDES SOCIALES (plantillas PNG)
# ===========================================================================
def piezas_redes():
    """Devuelve [(nombre, html, ancho, alto)] para renderizar como PNG."""
    def base(ancho, alto, contenido, fondo="#123C4C"):
        return envolver("kit", """<div id="lienzo" style="width:%dpx; height:%dpx;
             background:%s; position:relative; overflow:hidden; display:flex;
             flex-direction:column; justify-content:space-between; padding:%dpx;
             color:#EAF6F9">%s</div>""" % (ancho, alto, fondo, int(ancho * 0.075), contenido),
             css_extra="@page{margin:0} body{margin:0}")

    def marco(ancho, alto, hueco_txt, titular, escala):
        return """
        <div style="display:flex; justify-content:space-between; align-items:flex-start">
          <div class="logo blanco" style="font-size:%dpx"><span>KM0</span><span class="w"
            style="color:#EED8AE; font-size:%dpx">week</span></div>
          <div style="text-align:right; font-size:%dpx; color:#EED8AE; font-weight:800;
                      letter-spacing:.14em; line-height:1.5">%s</div>
        </div>
        <div style="flex:1; margin:%dpx 0; border:%dpx dashed rgba(238,216,174,.55);
             border-radius:%dpx; display:flex; align-items:center; justify-content:center;
             text-align:center">
          <div style="color:rgba(238,216,174,.8); font-size:%dpx; line-height:1.5;
               max-width:70%%">%s</div>
        </div>
        <div>
          <div style="font-family:'Lora',serif; font-weight:600; color:#fff; font-size:%dpx;
               line-height:1.1">%s</div>
          <div style="display:flex; justify-content:space-between; align-items:flex-end;
               margin-top:%dpx; font-size:%dpx; color:rgba(234,246,249,.85)">
            <span>%s<br>Reserva directa · %s</span>
            <span style="font-family:'Caveat Brush',cursive; color:#EED8AE; font-size:%dpx">
              Descubre lo cerca</span>
          </div>
        </div>""" % (
            int(46 * escala), int(50 * escala), int(15 * escala),
            esc(FECHAS).replace(" – ", "<br>"),
            int(28 * escala), max(2, int(3 * escala)), int(24 * escala), int(19 * escala),
            hueco_txt, int(52 * escala), titular,
            int(22 * escala), int(17 * escala), esc(FECHAS), WEB, int(30 * escala))

    def relleno(ancho, alto, titular, sub, escala, fondo):
        return """
        <div class="logo blanco" style="font-size:%dpx"><span>KM0</span><span class="w"
          style="color:#EED8AE; font-size:%dpx">week</span></div>
        <div>
          <div style="font-family:'Lora',serif; font-weight:600; color:#fff; font-size:%dpx;
               line-height:1.08; max-width:88%%">%s</div>
          <div style="color:#EAF6F9; font-size:%dpx; margin-top:%dpx; max-width:80%%;
               line-height:1.45">%s</div>
        </div>
        <div style="display:flex; justify-content:space-between; align-items:flex-end;
             border-top:1px solid rgba(234,246,249,.28); padding-top:%dpx; font-size:%dpx;
             color:rgba(234,246,249,.85)">
          <span><b style="color:#EED8AE">%s</b><br>%s</span>
          <span style="font-family:'Caveat Brush',cursive; color:#EED8AE; font-size:%dpx">
            vive lo nuestro</span>
        </div>""" % (
            int(46 * escala), int(50 * escala), int(60 * escala), titular,
            int(20 * escala), int(18 * escala), sub,
            int(20 * escala), int(16 * escala), esc(FECHAS), WEB, int(30 * escala))

    return [
        ("plantilla-story-1080x1920", base(1080, 1920, marco(
            1080, 1920, "Tu foto aquí<br><span style='font-size:.7em'>1080 × 1920 px</span>",
            "Aquí somos<br>Km 0", 1.7)), 1080, 1920),
        ("plantilla-cuadrado-1080x1080", base(1080, 1080, marco(
            1080, 1080, "Tu foto aquí<br><span style='font-size:.7em'>1080 × 1080 px</span>",
            "Aquí somos Km 0", 1.35)), 1080, 1080),
        ("plantilla-banner-1200x628", base(1200, 628, marco(
            1200, 628, "Tu foto aquí · 1200 × 628 px", "Aquí somos Km 0", 0.95)), 1200, 628),
        ("ejemplo-story-1080x1920", base(1080, 1920, relleno(
            1080, 1920, "Esta vez,<br>el huésped<br>eres tú",
            "Del %s abrimos las puertas a nuestros vecinos. Pregúntanos por la tarifa Km0."
            % esc(FECHAS), 1.7, "#123C4C")), 1080, 1920),
        ("ejemplo-cuadrado-1080x1080", base(1080, 1080, relleno(
            1080, 1080, "A veinte minutos<br>de tu casa",
            "Tres fines de semana de noviembre para dormir donde duermen los que vienen de "
            "fuera.", 1.35, "#14647D"), "#14647D"), 1080, 1080),
        ("ejemplo-banner-1200x628", base(1200, 628, relleno(
            1200, 628, "Todo lo bueno está más cerca",
            "%d alojamientos · %d plazas para residentes · %s" % (len(ALOJ), CUPO_TOTAL, esc(FECHAS)),
            0.95, "#123C4C")), 1200, 628),
    ]


LEEME_REDES = """KIT DE REDES SOCIALES · HOSBEC Km0 Week %s
================================================================

Qué hay aquí
------------
  plantilla-story-1080x1920.png      Story y reel vertical (Instagram, Facebook)
  plantilla-cuadrado-1080x1080.png   Publicación cuadrada (Instagram, LinkedIn)
  plantilla-banner-1200x628.png      Banner horizontal (Facebook, X, web)

  ejemplo-*.png                      Los mismos formatos ya montados, listos para publicar
                                     tal cual si no quieres editar nada.

Cómo se usan las plantillas
---------------------------
El recuadro discontinuo marca dónde va TU foto. Ábrelas con cualquier editor
(Canva, Photoshop, incluso PowerPoint), coloca la imagen DEBAJO de la plantilla
y ajústala al hueco. El marco, el logotipo y las fechas ya están puestos.

Qué foto poner
--------------
  · De tu casa, no de banco de imágenes. Se nota, y la gracia es que se note.
  · Con luz de día y sin gente posando.
  · Mejor un rincón concreto que una vista general: la terraza, la mesa del
    desayuno, la puerta de entrada.
  · Nada de collages ni marcos de colores encima.

Textos que puedes usar
----------------------
  «Del %s abrimos las puertas a nuestros vecinos.»
  «Esta vez, el huésped eres tú.»
  «A veinte minutos de tu casa.»
  «Precio para residentes y reserva directa. Sin intermediarios.»
  «Pregúntanos por el Pasaporte Km0.»

Etiquetas
---------
  #Km0Week #ComunitatValenciana #TurismeDeProximitat #HOSBEC
  Y la de tu municipio.

Colores de la marca
-------------------
  Azul Mediterráneo  #1EA4C6      Arena cálida    #EED8AE
  Azul tinta         #123C4C      Terracota       #D9794D
  Verde wellness     #8CB26F      Blanco roto     #FAF5EC

Tipografías (licencia libre, se pueden instalar sin coste)
  Titulares: Lora Semibold · Texto: Montserrat Medium
  Acentos escritos a mano: Caveat Brush

Lo que no se hace
-----------------
  · Cambiar los colores del logotipo.
  · Estirar o inclinar las plantillas: escálalas en proporción.
  · Tapar las fechas o la dirección de la web.
  · Prometer descuentos concretos que no estén en tu ficha.

Dudas: %s · %s
%s
"""


# ===========================================================================
# 11 · LISTADO DE ALOJAMIENTOS (hoja de cálculo)
# ===========================================================================
def hacer_xlsx(ruta):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Alojamientos"

    cols = [
        ("Nombre", 34), ("Tipo", 15), ("Categoría", 10), ("Municipio", 22),
        ("Provincia", 12), ("Plazas comprometidas", 12), ("Precio desde (€)", 12),
        ("Precio habitual (€)", 13), ("Descuento (%)", 11), ("Unidad", 26),
        ("Teléfono", 15), ("Web", 30), ("Experiencias", 40), ("Oferta", 34),
    ]
    ws.append([c[0] for c in cols])

    tinta = PatternFill("solid", fgColor="123C4C")
    blanco = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    normal = Font(name="Arial", size=10)
    borde = Border(bottom=Side(style="thin", color="E4E9EB"))

    for i, (t, an) in enumerate(cols, start=1):
        c = ws.cell(row=1, column=i)
        c.fill, c.font = tinta, blanco
        c.alignment = Alignment(vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(i)].width = an
    ws.row_dimensions[1].height = 30
    ws.freeze_panes = "A2"

    for a in sorted(ALOJ, key=lambda x: (x["provincia"], x["destino"], x["nombre"])):
        o = a.get("oferta") or {}
        ws.append([
            a["nombre"], tipo_es(a["tipo"]), a.get("categoria") or "", a["destino"],
            a["provincia"], a.get("cupo") or 0,
            o.get("precioDesde") or "", o.get("precioOriginal") or "", o.get("dto") or "",
            L(o.get("unidad")), a.get("telefono") or "", a.get("web") or "",
            ", ".join(exp_es(e) for e in a.get("experiencias") or []),
            L(o.get("titulo")),
        ])

    fin = ws.max_row
    for fila in ws.iter_rows(min_row=2, max_row=fin):
        for c in fila:
            c.font, c.border = normal, borde
            c.alignment = Alignment(vertical="top", wrap_text=(c.column in (13, 14)))
    for r in range(2, fin + 1):
        ws.cell(row=r, column=6).number_format = "#,##0"
        for col in (7, 8):
            ws.cell(row=r, column=col).number_format = '#,##0 "€"'
        ws.cell(row=r, column=9).number_format = '0"%"'

    # Totales con fórmulas, no con números calculados en Python
    t = fin + 2
    ws.cell(row=t, column=1, value="TOTALES").font = Font(name="Arial", bold=True, size=10)
    ws.cell(row=t, column=5, value="Alojamientos:").font = Font(name="Arial", bold=True, size=10)
    ws.cell(row=t, column=6, value="=COUNTA(A2:A%d)" % fin).font = Font(name="Arial", bold=True, size=10)
    ws.cell(row=t + 1, column=5, value="Plazas:").font = Font(name="Arial", bold=True, size=10)
    ws.cell(row=t + 1, column=6, value="=SUM(F2:F%d)" % fin).font = Font(name="Arial", bold=True, size=10)
    ws.cell(row=t + 1, column=6).number_format = "#,##0"
    ws.cell(row=t + 2, column=5, value="Dto. medio:").font = Font(name="Arial", bold=True, size=10)
    ws.cell(row=t + 2, column=6, value="=IFERROR(ROUND(AVERAGE(I2:I%d),0),0)" % fin).font = Font(name="Arial", bold=True, size=10)
    ws.cell(row=t + 2, column=6).number_format = '0"%"'

    # ---- resumen por provincia, también con fórmulas ----
    r2 = wb.create_sheet("Resumen")
    r2.append(["Provincia", "Alojamientos", "Plazas comprometidas"])
    for i in (1, 2, 3):
        c = r2.cell(row=1, column=i)
        c.fill, c.font = tinta, blanco
    for i, p in enumerate(PROVINCIAS, start=2):
        r2.cell(row=i, column=1, value=p).font = normal
        r2.cell(row=i, column=2, value="=COUNTIF(Alojamientos!E:E,A%d)" % i).font = normal
        r2.cell(row=i, column=3, value="=SUMIF(Alojamientos!E:E,A%d,Alojamientos!F:F)" % i).font = normal
    r2.cell(row=5, column=1, value="Total").font = Font(name="Arial", bold=True, size=10)
    r2.cell(row=5, column=2, value="=SUM(B2:B4)").font = Font(name="Arial", bold=True, size=10)
    r2.cell(row=5, column=3, value="=SUM(C2:C4)").font = Font(name="Arial", bold=True, size=10)
    for col, an in (("A", 20), ("B", 16), ("C", 22)):
        r2.column_dimensions[col].width = an

    r2.cell(row=7, column=1, value="Fuente: assets/js/data-alojamientos.js del sitio de la "
            "Km0 Week. Generado el %s." % HOY.strftime("%d/%m/%Y")).font = Font(
                name="Arial", size=9, italic=True, color="7D8F98")
    r2.cell(row=8, column=1, value="El cupo es el número de plazas que cada alojamiento "
            "compromete públicamente para residentes.").font = Font(
                name="Arial", size=9, italic=True, color="7D8F98")

    wb.save(ruta)


# ===========================================================================
# 12 · TEXTOS PARA TU WEB Y TU CORREO (DOCX, vía docx-js)
# ===========================================================================
TEXTOS = [
    ("Un párrafo para tu web", [
        ("es", "Somos alojamiento adherido a la HOSBEC Km0 Week. Del %s abrimos las puertas a "
               "quienes viven en la Comunitat Valenciana con una tarifa pensada para ellos y un "
               "cupo de plazas reservado. Se reserva directamente con nosotros, sin "
               "intermediarios, acreditando residencia en la Comunitat al llegar." % FECHAS),
        ("va", "Som allotjament adherit a la HOSBEC Km0 Week. Del %s obrim les portes a qui viu "
               "a la Comunitat Valenciana amb una tarifa pensada per a ells i un cupo de places "
               "reservat. Es reserva directament amb nosaltres, sense intermediaris, acreditant "
               "residència a la Comunitat en arribar." % CFG["fechasTexto"]["va"]),
    ]),
    ("Texto corto para la portada", [
        ("es", "Esta vez, el huésped eres tú. Km0 Week, del %s: tarifa para residentes y reserva "
               "directa." % FECHAS),
        ("va", "Aquesta vegada, l'hoste eres tu. Km0 Week, del %s: tarifa per a residents i "
               "reserva directa." % CFG["fechasTexto"]["va"]),
    ]),
    ("Correo a tu base de clientes", [
        ("es", "Asunto: Esta vez el huésped eres tú\n\nHola:\n\nDel %s participamos en la Km0 "
               "Week, la iniciativa de HOSBEC para que quienes viven en la Comunitat Valenciana "
               "redescubran los alojamientos de su propio territorio.\n\nDurante esos días "
               "tenemos una tarifa específica para residentes, con un cupo de plazas reservado y "
               "condiciones a la vista. Se reserva directamente con nosotros: ni comisiones ni "
               "intermediarios.\n\nAdemás hay un programa de actividades abiertas a todo el "
               "mundo, sin necesidad de alojarse, y un pasaporte que suma sellos y entra en el "
               "sorteo de diez estancias para 2027.\n\nTe esperamos.\n" % FECHAS),
        ("va", "Assumpte: Aquesta vegada l'hoste eres tu\n\nHola:\n\nDel %s participem en la Km0 "
               "Week, la iniciativa d'HOSBEC perquè qui viu a la Comunitat Valenciana "
               "redescobrisca els allotjaments del seu propi territori.\n\nDurant aqueixos dies "
               "tenim una tarifa específica per a residents, amb un cupo de places reservat i "
               "condicions a la vista. Es reserva directament amb nosaltres.\n\nT'esperem.\n"
               % CFG["fechasTexto"]["va"]),
    ]),
    ("Respuesta cuando pregunten por teléfono", [
        ("es", "«Sí, participamos en la Km0 Week. Es una tarifa solo para gente que viva en la "
               "Comunitat, del %s. Se reserva con nosotros directamente y solo hace falta traer "
               "el DNI o algo con domicilio en la Comunitat al llegar. Tenemos un cupo de plazas "
               "reservado para eso, así que cuanto antes mejor.»" % FECHAS),
        ("va", "«Sí, participem en la Km0 Week. És una tarifa només per a gent que visca a la "
               "Comunitat, del %s. Es reserva amb nosaltres directament i només cal portar el "
               "DNI o alguna cosa amb domicili a la Comunitat en arribar.»"
               % CFG["fechasTexto"]["va"]),
    ]),
    ("Pie para tu firma de correo", [
        ("es", "Alojamiento adherido a la HOSBEC Km0 Week · %s · %s" % (FECHAS, WEB)),
        ("va", "Allotjament adherit a la HOSBEC Km0 Week · %s · %s"
               % (CFG["fechasTexto"]["va"], WEB)),
    ]),
]



def hacer_docx(ruta):
    """Textos listos para copiar, en un Word con estilos. Se monta con docx-js."""
    datos = {"salida": ruta, "fechas": FECHAS, "web": WEB, "email": EMAIL, "tel": TEL,
             "hosbec": HOSBEC, "bloques": [
                 {"titulo": t, "versiones": [{"idioma": i, "texto": x} for i, x in vs]}
                 for t, vs in TEXTOS]}
    js = os.path.join(TMP, "docx.js")
    io.open(js, "w", encoding="utf-8").write(DOCX_JS)
    manif = os.path.join(TMP, "docx.json")
    io.open(manif, "w", encoding="utf-8").write(json.dumps(datos))
    subprocess.run(["node", js, manif], check=True)


DOCX_JS = r"""
const fs = require('fs');
const { Document, Packer, Paragraph, TextRun, HeadingLevel, AlignmentType,
        BorderStyle, ShadingType, Table, TableRow, TableCell, WidthType } = require('docx');
const d = JSON.parse(fs.readFileSync(process.argv[2], 'utf8'));

const AZUL = '123C4C', MAR = '1EA4C6', SUAVE = '7D8F98';
const hijos = [];

hijos.push(new Paragraph({ children: [
  new TextRun({ text: 'HOSBEC Km0 Week', bold: true, size: 44, color: AZUL, font: 'Georgia' })
]}));
hijos.push(new Paragraph({ spacing: { after: 120 }, children: [
  new TextRun({ text: 'Textos para tu web y tu correo', size: 30, color: MAR, font: 'Calibri' })
]}));
hijos.push(new Paragraph({ spacing: { after: 300 },
  border: { bottom: { style: BorderStyle.SINGLE, size: 8, color: 'E4E9EB' } },
  children: [ new TextRun({ text: d.fechas + '  ·  ' + d.web, size: 20, color: SUAVE,
                            font: 'Calibri' }) ]}));

hijos.push(new Paragraph({ spacing: { after: 260 }, children: [
  new TextRun({ text: 'Copia lo que necesites y adáptalo. Están en castellano y en valencià: '
    + 'si publicas en los dos idiomas, usa las dos versiones; si solo en uno, elige y no '
    + 'mezcles. Lo único que no conviene cambiar son las fechas, la condición de residencia '
    + 'y que la reserva es directa.', size: 21, font: 'Calibri', color: '4C6672' })
]}));

d.bloques.forEach((b, i) => {
  hijos.push(new Paragraph({ heading: HeadingLevel.HEADING_1,
    spacing: { before: i ? 360 : 120, after: 140 }, children: [
      new TextRun({ text: b.titulo, bold: true, size: 26, color: AZUL, font: 'Georgia' })
    ]}));
  b.versiones.forEach(v => {
    hijos.push(new Paragraph({ spacing: { before: 100, after: 60 }, children: [
      new TextRun({ text: v.idioma === 'va' ? 'VALENCIÀ' : 'CASTELLANO',
        bold: true, size: 15, color: v.idioma === 'va' ? '6E9553' : MAR,
        font: 'Calibri', characterSpacing: 30 })
    ]}));
    v.texto.split('\n\n').forEach(par => {
      hijos.push(new Paragraph({
        spacing: { after: 110 }, indent: { left: 220 },
        border: { left: { style: BorderStyle.SINGLE, size: 12,
                          color: v.idioma === 'va' ? 'DCE8CF' : 'CDE6F0', space: 10 } },
        children: [ new TextRun({ text: par.replace(/\n/g, '  '), size: 21,
                                  font: 'Calibri', color: '123C4C' }) ]}));
    });
  });
});

hijos.push(new Paragraph({ spacing: { before: 420, after: 100 },
  border: { top: { style: BorderStyle.SINGLE, size: 8, color: 'E4E9EB' } },
  children: [ new TextRun({ text: 'HOSBEC · ' + d.hosbec, size: 17, color: SUAVE,
                            font: 'Calibri' }) ]}));
hijos.push(new Paragraph({ children: [
  new TextRun({ text: d.email + '  ·  ' + d.tel, size: 17, color: SUAVE, font: 'Calibri' })
]}));

const doc = new Document({ creator: 'HOSBEC', title: 'Textos Km0 Week',
  sections: [{ properties: { page: { margin: { top: 1200, bottom: 1100,
                                               left: 1200, right: 1200 } } },
               children: hijos }] });
Packer.toBuffer(doc).then(buf => fs.writeFileSync(d.salida, buf));
"""


# ===========================================================================
# 13 · Motor: HTML → PDF/PNG con Chromium (para conservar las tipografías)
# ===========================================================================
RENDER_JS = r"""
const { chromium } = require(process.env.PW || 'playwright');
const fs = require('fs');
(async () => {
  const trabajos = JSON.parse(fs.readFileSync(process.argv[2], 'utf8'));
  const b = await chromium.launch(
    process.env.PW_CHROME ? { executablePath: process.env.PW_CHROME } : {});
  for (const t of trabajos) {
    const p = await b.newPage(t.ancho ? { viewport: { width: t.ancho, height: t.alto } } : {});
    const errs = [];
    p.on('pageerror', e => errs.push(e.message));
    await p.goto('file://' + t.html, { waitUntil: 'networkidle' });
    await p.evaluate(() => document.fonts.ready);
    await p.waitForTimeout(250);
    if (t.tipo === 'png') {
      const el = await p.$('#lienzo');
      await (el || p).screenshot({ path: t.salida });
    } else if (t.pieRepetido) {
      // Documentos largos: el pie va en el margen inferior de CADA página,
      // que es lo que hace Chromium con displayHeaderFooter.
      await p.pdf({ path: t.salida, printBackground: true, format: t.formato || 'A4',
        displayHeaderFooter: true, headerTemplate: '<div></div>',
        footerTemplate: '<div style="width:100%; font-size:7pt; font-family:sans-serif;' +
          ' color:#7D8F98; padding:0 15mm; display:flex; justify-content:space-between">' +
          '<span>' + t.pieRepetido + '</span>' +
          '<span>pág. <span class="pageNumber"></span> de <span class="totalPages"></span>' +
          '</span></div>',
        margin: { top: '14mm', bottom: '17mm', left: '15mm', right: '15mm' } });
    } else {
      await p.pdf({ path: t.salida, printBackground: true,
                    format: t.formato || 'A4', landscape: !!t.apaisado,
                    preferCSSPageSize: true });
    }
    if (errs.length) console.log('  ! ' + t.salida + ': ' + errs[0]);
    await p.close();
  }
  await b.close();
})();
"""


def renderizar(trabajos):
    js = os.path.join(TMP, "render.js")
    io.open(js, "w", encoding="utf-8").write(RENDER_JS)
    manifiesto = os.path.join(TMP, "trabajos.json")
    io.open(manifiesto, "w", encoding="utf-8").write(json.dumps(trabajos))
    env = dict(os.environ)
    # Dónde está el paquete playwright. Si no viene por entorno, se prueba la
    # instalación global de Cowork; si tampoco está, se deja que Node lo
    # resuelva solo (con "npm install playwright" en la raíz del repositorio,
    # que es lo que hace la publicación automática de GitHub).
    if "PW" not in env:
        glob = "/home/claude/.npm-global/lib/node_modules/playwright"
        env["PW"] = glob if os.path.isdir(glob) else "playwright"
    subprocess.run(["node", js, manifiesto], check=True, env=env)


def escribir_html(nombre, html):
    ruta = os.path.join(HTML, nombre + ".html")
    io.open(ruta, "w", encoding="utf-8").write(html)
    return ruta


def kb(ruta):
    n = os.path.getsize(ruta)
    return "%.1f MB" % (n / 1048576.0) if n >= 1048576 else "%d KB" % max(1, round(n / 1024.0))


# ===========================================================================
# 14 · Construir todo
# ===========================================================================
PIES_REPETIDOS = {
    "programa-actividades": "Programa de actividades · HOSBEC Km0 Week 2026",
    "bases-sorteo": "Bases del sorteo · Km0 Week 2026 · Borrador pendiente de validación jurídica",
}


def main():
    print("Km0 Week · generando los descargables\n")
    trabajos = []

    docs = [
        ("pasaporte-km0",        doc_pasaporte(),  "A4", True),
        ("programa-actividades", doc_programa(),   "A4", False),
        ("bases-sorteo",         doc_bases(),      "A4", False),
        ("carteleria",           doc_carteleria(), "A4", False),
        ("manual-de-marca",      doc_manual(),     "A4", False),
        ("sello-pasaporte",      doc_sello(),      "A4", False),
        ("guia-recepcion",       doc_guia(),       "A4", False),
        ("dossier-prensa",       doc_dossier(),    "A4", False),
        ("nota-prensa-presentacion", doc_nota(),   "A4", False),
    ]
    for nombre, html, formato, apaisado in docs:
        t = {"html": escribir_html(nombre, html),
             "salida": os.path.join(SALIDA, nombre + ".pdf"),
             "tipo": "pdf", "formato": formato, "apaisado": apaisado}
        if nombre in PIES_REPETIDOS:
            t["pieRepetido"] = PIES_REPETIDOS[nombre]
        trabajos.append(t)

    redes = piezas_redes()
    for nombre, html, an, al in redes:
        trabajos.append({"html": escribir_html("redes-" + nombre, html),
                         "salida": os.path.join(TMP, nombre + ".png"),
                         "tipo": "png", "ancho": an, "alto": al})

    # el logotipo, en PNG a tres tamaños y sobre los dos fondos
    for fondo, sufijo, color, colorw in (("#FFFFFF", "claro", "#123C4C", "#1EA4C6"),
                                         ("#123C4C", "oscuro", "#FFFFFF", "#EED8AE")):
        for an in (400, 800, 1600):
            al = int(an * 0.32)
            html = envolver("logo", """<div id="lienzo" style="width:%dpx; height:%dpx;
                 background:%s; display:flex; align-items:center; justify-content:center">
                 <span class="logo" style="font-size:%dpx; color:%s"><span>KM0</span><span
                 class="w" style="font-size:%dpx; color:%s">week</span></span></div>""" % (
                an, al, fondo, int(an * 0.19), color, int(an * 0.21), colorw),
                css_extra="@page{margin:0} body{margin:0}")
            n = "km0week-logo-%s-%dpx" % (sufijo, an)
            trabajos.append({"html": escribir_html("logo-" + n, html),
                             "salida": os.path.join(TMP, n + ".png"),
                             "tipo": "png", "ancho": an, "alto": al})

    print("· imprimiendo %d piezas con Chromium…" % len(trabajos))
    renderizar(trabajos)

    # ---------------------------------------------------------------- xlsx --
    print("· hoja de cálculo de alojamientos…")
    hacer_xlsx(os.path.join(SALIDA, "alojamientos-adheridos.xlsx"))

    # ---------------------------------------------------------------- docx --
    print("· documento de textos…")
    hacer_docx(os.path.join(SALIDA, "textos-para-tu-web.docx"))

    # ----------------------------------------------------------------- zips --
    print("· empaquetando…")
    zredes = os.path.join(SALIDA, "kit-redes-sociales.zip")
    with zipfile.ZipFile(zredes, "w", zipfile.ZIP_DEFLATED) as z:
        for nombre, _, _, _ in redes:
            z.write(os.path.join(TMP, nombre + ".png"), nombre + ".png")
        z.writestr("LEEME.txt", LEEME_REDES % (
            CFG["edicion"], FECHAS, EMAIL, TEL, ""))

    zlogos = os.path.join(SALIDA, "logotipos-y-marca.zip")
    with zipfile.ZipFile(zlogos, "w", zipfile.ZIP_DEFLATED) as z:
        for f in ("emblema.svg", "favicon.svg"):
            o = os.path.join(RAIZ, "assets", "img", f)
            if os.path.exists(o):
                z.write(o, "vectorial/" + f)
        for fondo in ("claro", "oscuro"):
            for an in (400, 800, 1600):
                n = "km0week-logo-%s-%dpx.png" % (fondo, an)
                z.write(os.path.join(TMP, n), "png/" + n)
        for f in sorted(os.listdir(FUENTES)):
            z.write(os.path.join(FUENTES, f), "tipografias/" + f)
        z.write(os.path.join(SALIDA, "manual-de-marca.pdf"), "manual-de-marca.pdf")
        z.writestr("LEEME.txt", LEEME_LOGOS)

    hacer_banco(os.path.join(SALIDA, "banco-imagenes.zip"))

    print("\nHecho. En descargas/:")
    for f in sorted(os.listdir(SALIDA)):
        print("   %-34s %s" % (f, kb(os.path.join(SALIDA, f))))


# ---------------------------------------------------------------------------
# Banco de imágenes: no genera nada, empaqueta lo que ya usa la web. Se
# construye a partir de assets/img/foto, así que crece solo a medida que se
# sustituyan las ilustraciones provisionales por fotografía real.
# ---------------------------------------------------------------------------
GRUPOS_BANCO = [
    ("prov-", "01-destinos",      "Los tres destinos provinciales"),
    ("alo-",  "02-alojamientos",  "Alojamientos adheridos"),
    ("tema-", "03-experiencias",  "Las experiencias del catálogo"),
    ("idea-", "04-la-iniciativa", "Imágenes de la iniciativa"),
    ("not-",  "05-noticias",      "Imágenes de las noticias publicadas"),
    ("cab-",  "06-cabeceras",     "Cabeceras de cada página de la web"),
]


def hacer_banco(destino):
    foto = os.path.join(RAIZ, "assets", "img", "foto")
    if not os.path.isdir(foto):
        return
    archivos = sorted(f for f in os.listdir(foto) if f.lower().endswith((".webp", ".jpg", ".png")))
    reparto = {c: [] for _, c, _ in GRUPOS_BANCO}
    sueltos = []
    for f in archivos:
        for pre, carpeta, _ in GRUPOS_BANCO:
            if f.startswith(pre):
                reparto[carpeta].append(f)
                break
        else:
            sueltos.append(f)

    L = [LEEME_BANCO % (HOY.strftime("%d/%m/%Y"), len(archivos), WEB),
         "CONTENIDO DEL ZIP", "-" * 60]
    for _, carpeta, titulo in GRUPOS_BANCO:
        if not reparto[carpeta]:
            continue
        L.append("")
        L.append("%s/  —  %s (%d)" % (carpeta, titulo, len(reparto[carpeta])))
        L += ["    · " + f for f in reparto[carpeta]]
    if sueltos:
        L += ["", "otras/ (%d)" % len(sueltos)] + ["    · " + f for f in sueltos]
    L += ["", "CONTACTO", "-" * 60,
          "%s · %s" % (EMAIL, TEL), HOSBEC, ""]

    with zipfile.ZipFile(destino, "w", zipfile.ZIP_DEFLATED) as z:
        z.writestr("LEEME-creditos-y-uso.txt", "\r\n".join(L))
        for _, carpeta, _ in GRUPOS_BANCO:
            for f in reparto[carpeta]:
                z.write(os.path.join(foto, f), "%s/%s" % (carpeta, f))
        for f in sueltos:
            z.write(os.path.join(foto, f), "otras/" + f)


LEEME_BANCO = """BANCO DE IMÁGENES · HOSBEC Km0 Week 2026
============================================================

Generado el %s.
%d imágenes en formato WebP, las mismas que usa la web de la iniciativa
en %s

AVISO IMPORTANTE SOBRE ESTAS IMÁGENES
------------------------------------------------------------
Son las imágenes provisionales de la web previa al lanzamiento. Ilustran
los destinos, los tipos de alojamiento y las experiencias, pero NO son
fotografías reales de los establecimientos adheridos ni de las actividades
del programa. Se irán sustituyendo por material fotográfico propio a medida
que la edición se cierre.

Hasta entonces: úsalas para maquetar, no las publiques como si fueran una
foto documental de un hotel concreto. Si necesitas material real de un
alojamiento, escríbenos y se lo pedimos al establecimiento.

CONDICIONES DE USO
------------------------------------------------------------
Uso editorial y promocional libre para hablar de la Km0 Week, citando
«HOSBEC · Km0 Week 2026». No se pueden revender ni ceder a terceros para
usos ajenos a la iniciativa.

FORMATO
------------------------------------------------------------
WebP, proporción 3:2, optimizadas para web. Si tu maquetador o tu imprenta
no acepta WebP, escríbenos y te las mandamos en JPG a resolución de
impresión.

"""


LEEME_LOGOS = """LOGOTIPOS Y MARCA · HOSBEC Km0 Week
====================================================

  vectorial/     El emblema y el favicon en SVG (escalables sin pérdida)
  png/           El logotipo completo a tres tamaños, sobre fondo claro y oscuro
  tipografias/   Las tres tipografías de la marca, en woff2
  manual-de-marca.pdf   Las reglas de uso completas

Sobre el logotipo
-----------------
El logotipo es un lockup de dos piezas: «KM0» compuesto en Montserrat
ExtraBold y «week» en Caveat Brush. Los PNG de esta carpeta están generados
con esas tipografías exactas y a los tamaños de uso más habituales.

Si necesitas el lockup en vectorial para una pieza de gran formato, se
compone en cualquier programa de diseño con las dos tipografías incluidas
aquí (ambas con licencia SIL Open Font License, uso libre y gratuito) y se
trazan las letras. Las proporciones y el área de respeto están en el manual.

Uso sobre fondo claro:  KM0 en azul tinta #123C4C, week en #1EA4C6
Uso sobre fondo oscuro: KM0 en blanco,     week en arena  #EED8AE

Tamaño mínimo: 110 px de ancho en pantalla, 28 mm impreso.

Lo que no se hace: cambiar los colores, sustituir las tipografías, deformar,
inclinar, añadir sombras o contornos, o meterlo dentro de una caja de color.

Consultas: km0week@hosbec.com
"""


if __name__ == "__main__":
    main()
